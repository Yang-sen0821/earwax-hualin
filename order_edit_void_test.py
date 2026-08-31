# -*- coding: utf-8 -*-
"""CR-4 驗收：訂單編輯 / 作廢（軟刪除 + SALE_REVERSAL + audit）— 2026-08-31。

throwaway sqlite 放 %TEMP%，不污染交付庫；env 在 import 任何 app 模組前先設定。
驗證項目：
  A 未出貨單編輯：盒 3→5、再 5→2 → 兩池餘量與 movement 筆數正確；連續編輯兩次
  B 第二次編輯缺貨 → 整張 rollback，餘量 / 品項 / movement 數不變
  C 作廢未出貨單 → 全額回補 normal，不碰 reserved；重複 reverse 被拒；重複作廢被拒；原因必填
  D 作廢後 /orders/、/reports/sales、Excel、客戶明細不含該單；?show_voided=1 灰字「已作廢」
  E 已出貨單：staff 作廢 403；staff 編輯只能改收件/運費/備註（品項鎖定不重扣）；owner 可作廢且不回補；paid→refunded
  F 手機：狀態 POST 拒絕非閉集值（cancelled / 亂字串）；手機作廢同規則
  G 出貨下拉無 cancelled；桌面 update_shipping 轉 shipped 補建 shipments 列
  H audit_logs：order_edit / order_void 各有紀錄；已售統計 = SALE − 回補淨額；Excel 異動分頁含「銷售回補」
"""
import os
import sys
import io
import json
import tempfile

try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except Exception:
    pass

_TMP_DB = os.path.join(tempfile.gettempdir(), "flora_court_editvoid_test.db")
if os.path.exists(_TMP_DB):
    os.remove(_TMP_DB)
os.environ["DATABASE_URL"] = "sqlite:///" + _TMP_DB.replace("\\", "/")
os.environ["ENABLE_MOBILE"] = "true"
os.environ["SECRET_KEY"] = "test-secret"

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import init_db  # noqa: E402
from db import (  # noqa: E402
    get_session, Order, OrderItem, InventoryBalance, InventoryMovement,
    Product, Shipment, AuditLog, Customer,
)
import inventory_service  # noqa: E402
from app import create_app  # noqa: E402

PASS = []
FAIL = []


def check(name, cond, detail=""):
    (PASS if cond else FAIL).append(name)
    print(f"[{'PASS' if cond else 'FAIL'}] {name}" + (f" — {detail}" if detail else ""))


def login(client, username, password):
    return client.post("/login", data={"username": username, "password": password},
                       follow_redirects=True)


def fresh():
    s = get_session()
    s.expire_all()
    return s


def bal(pid, pool, cat="normal"):
    db = fresh()
    b = db.query(InventoryBalance).filter_by(
        product_id=pid, inventory_pool=pool, stock_category=cat).first()
    q = b.qty if b else None
    db.close()
    return q


def last_order_id():
    db = fresh()
    o = db.query(Order).order_by(Order.id.desc()).first()
    oid, ono = o.id, o.order_no
    db.close()
    return oid, ono


def order(oid):
    db = fresh()
    o = db.query(Order).filter_by(id=oid).first()
    db.close()
    return o


def items(oid):
    db = fresh()
    rows = db.query(OrderItem).filter_by(order_id=oid).order_by(OrderItem.id).all()
    out = [(r.combo_code, r.qty, float(r.subtotal or 0)) for r in rows]
    db.close()
    return out


def mv_count(oid, mtype=None):
    db = fresh()
    q = db.query(InventoryMovement).filter_by(ref_type="order", ref_id=str(oid))
    if mtype:
        q = q.filter_by(movement_type=mtype)
    n = q.count()
    db.close()
    return n


def all_mv_count():
    db = fresh()
    n = db.query(InventoryMovement).count()
    db.close()
    return n


def audit_count(action, oid=None):
    db = fresh()
    q = db.query(AuditLog).filter_by(action=action)
    if oid is not None:
        q = q.filter_by(target_id=str(oid))
    n = q.count()
    db.close()
    return n


def audit_last(action, oid):
    db = fresh()
    a = (db.query(AuditLog).filter_by(action=action, target_id=str(oid))
         .order_by(AuditLog.id.desc()).first())
    d = json.loads(a.detail) if a else None
    db.close()
    return d


def shipments_count(oid):
    db = fresh()
    n = db.query(Shipment).filter_by(order_id=oid).count()
    db.close()
    return n


def main():
    init_db.create_all()
    s = get_session()
    try:
        init_db.seed(s)
    finally:
        s.close()

    app = create_app()
    app.config["TESTING"] = True
    oc = app.test_client()
    login(oc, "owner", "owner123")
    sc = app.test_client()
    login(sc, "staff", "staff123")

    db = get_session()
    pid = db.query(Product).filter_by(sku="FC-MASK-001").first().id
    db.close()
    box0 = bal(pid, "boxed")
    loose0 = bal(pid, "loose_piece")
    res0 = bal(pid, "boxed", "reserved")
    print(f"初始：盒裝normal={box0} 裸片normal={loose0} 盒裝reserved={res0}")

    # =====================================================================
    # A. 建單 盒×3 + 片×2 → 編輯 盒×5 → 編輯 盒×2/片×4
    # =====================================================================
    r = sc.post("/orders/new", data={
        "new_customer_name": "作廢測試客", "recipient_name": "作廢測試客",
        "item_product_id": [str(pid), str(pid)], "item_combo_code": ["BOX", "LOOSE"],
        "item_qty": ["3", "2"], "item_amount": ["900", "100"],
    }, follow_redirects=True)
    oid, ono = last_order_id()
    check("A 建單 盒×3 片×2", bal(pid, "boxed") == box0 - 3 and bal(pid, "loose_piece") == loose0 - 2)
    check("A 建單寫 2 筆 SALE", mv_count(oid, "SALE") == 2)

    g = sc.get(f"/orders/{oid}/edit")
    html = g.get_data(as_text=True)
    check("A 編輯頁 200 且預填品項 JSON", g.status_code == 200 and '"combo_code": "BOX"' in html
          and '"qty": 3' in html and "編輯訂單" in html)

    r = sc.post(f"/orders/{oid}/edit", data={
        "recipient_name": "作廢測試客-改", "discount": "100",
        "item_product_id": [str(pid), str(pid)], "item_combo_code": ["BOX", "LOOSE"],
        "item_qty": ["5", "2"], "item_amount": ["1500", "100"],
    }, follow_redirects=True)
    check("A 第一次編輯 200 有成功訊息", r.status_code == 200 and "訂單已更新" in r.get_data(as_text=True))
    check("A 盒 3→5：盒裝 normal = 初始−5", bal(pid, "boxed") == box0 - 5,
          f"now={bal(pid, 'boxed')} expect={box0-5}")
    check("A 片維持 2：裸片 = 初始−2", bal(pid, "loose_piece") == loose0 - 2)
    check("A movement：2 SALE_REVERSAL + 4 SALE", mv_count(oid, "SALE_REVERSAL") == 2
          and mv_count(oid, "SALE") == 4, f"rev={mv_count(oid, 'SALE_REVERSAL')} sale={mv_count(oid, 'SALE')}")
    o = order(oid)
    check("A total = 1600 − 100 = 1500；收件人已改", float(o.total_amount) == 1500.0
          and float(o.discount) == 100.0 and o.recipient_name == "作廢測試客-改", f"total={o.total_amount}")
    check("A items 重建為 盒5/片2", items(oid) == [("BOX", 5, 1500.0), ("LOOSE", 2, 100.0)], str(items(oid)))
    check("A audit order_edit 1 筆", audit_count("order_edit", oid) == 1)
    d = audit_last("order_edit", oid)
    check("A audit detail 含 before/after 與 reversal ids", d and d["before"]["items"][0]["qty"] == 3
          and d["after"]["items"][0]["qty"] == 5 and len(d["reversal_movement_ids"]) == 2)

    r = sc.post(f"/orders/{oid}/edit", data={
        "discount": "0",
        "item_product_id": [str(pid), str(pid)], "item_combo_code": ["BOX", "LOOSE"],
        "item_qty": ["2", "4"], "item_amount": ["600", "200"],
    }, follow_redirects=True)
    check("A 第二次編輯 盒 5→2 片 2→4：盒裝 = 初始−2", bal(pid, "boxed") == box0 - 2,
          f"now={bal(pid, 'boxed')}")
    check("A 第二次編輯 裸片 = 初始−4", bal(pid, "loose_piece") == loose0 - 4)
    check("A movement：4 SALE_REVERSAL + 6 SALE", mv_count(oid, "SALE_REVERSAL") == 4
          and mv_count(oid, "SALE") == 6)
    check("A total = 800", float(order(oid).total_amount) == 800.0)
    check("A audit order_edit 2 筆", audit_count("order_edit", oid) == 2)

    # =====================================================================
    # B. 第三次編輯缺貨 → 整張 rollback
    # =====================================================================
    mv_all = all_mv_count()
    r = sc.post(f"/orders/{oid}/edit", data={
        "discount": "0",
        "item_product_id": [str(pid), str(pid)], "item_combo_code": ["BOX", "LOOSE"],
        "item_qty": ["1", "999999"], "item_amount": ["300", "1"],
    }, follow_redirects=True)
    check("B 缺貨編輯有提示", "庫存不足" in r.get_data(as_text=True))
    check("B rollback 後盒裝不變（初始−2）", bal(pid, "boxed") == box0 - 2, f"now={bal(pid, 'boxed')}")
    check("B rollback 後裸片不變（初始−4）", bal(pid, "loose_piece") == loose0 - 4)
    check("B rollback 後 movement 總數不變", all_mv_count() == mv_all)
    check("B rollback 後品項不變 盒2/片4", items(oid) == [("BOX", 2, 600.0), ("LOOSE", 4, 200.0)])
    check("B rollback 後 total 不變 800", float(order(oid).total_amount) == 800.0)
    check("B audit 未增加", audit_count("order_edit", oid) == 2)

    # =====================================================================
    # C. 作廢未出貨單（staff）→ 全額回補；不碰 reserved；重複被拒
    # =====================================================================
    r = sc.post(f"/orders/{oid}/void", data={"reason": ""}, follow_redirects=True)
    check("C 無原因作廢被拒", "原因" in r.get_data(as_text=True) and order(oid).voided_at is None)
    r = sc.post(f"/orders/{oid}/void", data={"reason": "客戶取消"}, follow_redirects=True)
    check("C staff 作廢未出貨單成功", "已作廢" in r.get_data(as_text=True) and order(oid).voided_at is not None)
    check("C 盒裝回補到初始", bal(pid, "boxed") == box0, f"now={bal(pid, 'boxed')} expect={box0}")
    check("C 裸片回補到初始", bal(pid, "loose_piece") == loose0)
    check("C reserved 全程未動", bal(pid, "boxed", "reserved") == res0)
    check("C 作廢寫 2 筆 SALE_REVERSAL（共 6）", mv_count(oid, "SALE_REVERSAL") == 6)
    o = order(oid)
    check("C voided_by / void_reason 落地", o.voided_by is not None and o.void_reason == "客戶取消")
    check("C audit order_void 1 筆", audit_count("order_void", oid) == 1)
    d = audit_last("order_void", oid)
    check("C audit detail 含快照/total/movement ids", d and d["snapshot"]["total_amount"] == "800.00"
          and len(d["reversal_movement_ids"]) == 2 and d["stock_reversed"] is True)
    check("C items/payments 未刪（order_items 仍 2 列）", len(items(oid)) == 2)

    db = get_session()
    rv = inventory_service.reverse_sale(db, oid, operator="tester", reason="dup")
    db.rollback()
    db.close()
    check("C 重複 reverse_sale 被拒（已回補過）", rv.ok is False and "已回補過" in (rv.error or ""), rv.error)
    r = sc.post(f"/orders/{oid}/void", data={"reason": "again"}, follow_redirects=True)
    check("C 重複作廢被拒", "已作廢" in r.get_data(as_text=True) and audit_count("order_void", oid) == 1)
    check("C 作廢後盒裝仍 = 初始（無二次回補）", bal(pid, "boxed") == box0)
    r = sc.post(f"/orders/{oid}/edit", data={"item_product_id": str(pid), "item_combo_code": "BOX",
                                            "item_qty": "1", "item_amount": "1"}, follow_redirects=True)
    check("C 作廢單不可編輯", "不可編輯" in r.get_data(as_text=True) and bal(pid, "boxed") == box0)
    r = sc.post(f"/orders/{oid}/payment", data={"payment_status": "paid"}, follow_redirects=True)
    check("C 作廢單不可改付款", order(oid).payment_status == "unpaid")

    # =====================================================================
    # D. 統計排除：/orders/、/reports/sales、Excel、客戶明細
    # =====================================================================
    lst = oc.get("/orders/").get_data(as_text=True)
    check("D /orders/ 預設不含作廢單", ono not in lst)
    lst2 = oc.get("/orders/?show_voided=1").get_data(as_text=True)
    check("D /orders/?show_voided=1 含該單且標「已作廢」", ono in lst2 and "已作廢" in lst2
          and "line-through" in lst2)
    sales = oc.get("/reports/sales").get_data(as_text=True)
    check("D /reports/sales 200 且尚無銷售資料（唯一單已作廢）", "尚無銷售資料" in sales)
    xlsx = oc.get("/reports/export.xlsx")
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(xlsx.data))
    flat_o = [str(c.value) for row in wb["訂單"].iter_rows() for c in row if c.value is not None]
    check("D Excel 訂單分頁不含作廢單", not any(ono in v for v in flat_o))
    flat_m = [str(c.value) for row in wb["庫存異動"].iter_rows() for c in row if c.value is not None]
    check("H Excel 異動分頁含「銷售回補」", any("銷售回補" in v for v in flat_m))
    db = fresh()
    cid = db.query(Customer).filter_by(name="作廢測試客").first().id
    db.close()
    cdet = oc.get(f"/customers/{cid}").get_data(as_text=True)
    check("D 客戶明細不含作廢單", ono not in cdet and "尚無訂單" in cdet)
    idx = oc.get("/reports/").get_data(as_text=True)
    check("D 報表首頁 KPI 訂單總數 0", ">0<" in idx.replace(" ", ""), "")
    mlst = sc.get("/m/orders").get_data(as_text=True)
    check("D 手機列表預設不含作廢單", ono not in mlst)
    mlst2 = sc.get("/m/orders?voided=1").get_data(as_text=True)
    check("D 手機列表 ?voided=1 含作廢單標示", ono in mlst2 and "已作廢" in mlst2)
    check("D 手機待出貨不含作廢單", ono not in sc.get("/m/shipments").get_data(as_text=True))

    # =====================================================================
    # E. 已出貨單（桌面出貨補建 shipments）：staff 作廢 403；staff 編輯品項鎖定；owner 作廢不回補；paid→refunded
    # =====================================================================
    bag0 = None
    db = fresh()
    pkg = db.query(Product).filter_by(sku="PKG-SHARED-001").first()
    db.close()
    bag0 = bal(pkg.id, "paper_bag")
    sc.post("/orders/new", data={
        "item_product_id": str(pid), "item_combo_code": "BOX", "item_qty": "1", "item_amount": "300",
    }, follow_redirects=True)
    oid2, ono2 = last_order_id()
    box_after_o2 = bal(pid, "boxed")
    check("E 建單 盒×1", box_after_o2 == box0 - 1)
    det = oc.get(f"/orders/{oid2}").get_data(as_text=True)
    check("G 桌面出貨下拉無 cancelled", 'value="cancelled"' not in det and 'value="shipped"' in det)
    r = oc.post(f"/orders/{oid2}/shipping", data={"shipping_status": "cancelled"}, follow_redirects=True)
    check("G 桌面 POST cancelled 被拒", "無效的出貨狀態" in r.get_data(as_text=True)
          and order(oid2).shipping_status == "pending")
    r = oc.post(f"/orders/{oid2}/shipping", data={"shipping_status": "shipped", "carrier": "黑貓",
                                                  "tracking_no": "T123"}, follow_redirects=True)
    check("G 桌面轉 shipped 補建 shipments 列", shipments_count(oid2) == 1
          and order(oid2).shipping_status == "shipped")
    check("G 桌面出貨扣紙袋 1", bal(pkg.id, "paper_bag") == bag0 - 1)
    oc.post(f"/orders/{oid2}/payment", data={"payment_status": "paid"}, follow_redirects=True)
    check("E 付款改 paid", order(oid2).payment_status == "paid")

    r = sc.post(f"/orders/{oid2}/void", data={"reason": "staff 想作廢"})
    check("E 已出貨單 staff 作廢 → 403", r.status_code == 403, f"status={r.status_code}")
    check("E 403 後未作廢、庫存不動", order(oid2).voided_at is None and bal(pid, "boxed") == box_after_o2)

    g = sc.get(f"/orders/{oid2}/edit").get_data(as_text=True)
    check("E 已出貨單編輯頁顯示品項鎖定", "品項與折扣鎖定" in g and "var LOCKED = true" in g)
    r = sc.post(f"/orders/{oid2}/edit", data={
        "recipient_name": "改收件人", "note": "改備註", "shipping_fee": "99", "shipping_method": "post",
        "discount": "50",
        "item_product_id": str(pid), "item_combo_code": "BOX", "item_qty": "10", "item_amount": "3000",
    }, follow_redirects=True)
    o2 = order(oid2)
    check("E staff 編輯已出貨單：收件/備註/運費/方式已改", o2.recipient_name == "改收件人"
          and o2.note == "改備註" and float(o2.shipping_fee) == 99.0 and o2.shipping_method == "post")
    check("E 已出貨單品項/折扣/total 不變、庫存不重扣", items(oid2) == [("BOX", 1, 300.0)]
          and float(o2.discount or 0) == 0.0 and float(o2.total_amount) == 300.0
          and bal(pid, "boxed") == box_after_o2, str(items(oid2)))
    check("E 已出貨單 movement 無新增（無回補無重扣）", mv_count(oid2) == 1)
    d = audit_last("order_edit", oid2)
    check("E audit order_edit items_locked=true", d and d["items_locked"] is True)

    r = oc.post(f"/orders/{oid2}/void", data={"reason": "owner 作廢已出貨"}, follow_redirects=True)
    o2 = order(oid2)
    check("E owner 作廢已出貨單成功", o2.voided_at is not None and "不回補" in r.get_data(as_text=True))
    check("E 已出貨作廢不回補（盒裝仍 初始−1）", bal(pid, "boxed") == box_after_o2
          and mv_count(oid2, "SALE_REVERSAL") == 0)
    check("E paid → refunded", o2.payment_status == "refunded")
    check("E shipments 列保留", shipments_count(oid2) == 1)
    d = audit_last("order_void", oid2)
    check("E audit order_void stock_reversed=false, shipped=true", d and d["stock_reversed"] is False
          and d["shipped"] is True and d["payment_status_after"] == "refunded")

    # =====================================================================
    # F. 手機：狀態 POST 閉集驗證；手機作廢
    # =====================================================================
    sc.post("/m/orders/new", data={"product_id": str(pid), "combo_code": "LOOSE", "qty": "3",
                                   "amount": "150"}, follow_redirects=True)
    oid3, ono3 = last_order_id()
    loose_o3 = bal(pid, "loose_piece")
    mdet = sc.get(f"/m/orders/{oid3}").get_data(as_text=True)
    check("G 手機明細出貨下拉無 cancelled，且有作廢表單", 'value="cancelled"' not in mdet
          and f"/m/orders/{oid3}/void" in mdet)
    r = sc.post(f"/m/orders/{oid3}", data={"shipping_status": "cancelled"}, follow_redirects=True)
    check("F 手機 POST shipping_status=cancelled 被拒", "無效的出貨狀態值" in r.get_data(as_text=True)
          and order(oid3).shipping_status == "pending")
    r = sc.post(f"/m/orders/{oid3}", data={"payment_status": "bogus"}, follow_redirects=True)
    check("F 手機 POST payment_status 亂字串被拒", "無效的付款狀態值" in r.get_data(as_text=True)
          and order(oid3).payment_status == "unpaid")
    r = sc.post(f"/m/orders/{oid3}", data={"payment_status": "paid", "shipping_status": "pending"},
                follow_redirects=True)
    check("F 手機 POST 合法值可更新", order(oid3).payment_status == "paid")

    r = sc.post(f"/m/orders/{oid3}/void", data={"reason": ""}, follow_redirects=True)
    check("F 手機作廢無原因被拒", order(oid3).voided_at is None)
    r = sc.post(f"/m/orders/{oid3}/void", data={"reason": "手機作廢"}, follow_redirects=True)
    o3 = order(oid3)
    check("F 手機 staff 作廢未出貨單 → 回補 + paid→refunded", o3.voided_at is not None
          and bal(pid, "loose_piece") == loose_o3 + 3 and o3.payment_status == "refunded")
    check("F 手機作廢 audit", audit_count("order_void", oid3) == 1)
    mdet = sc.get(f"/m/orders/{oid3}").get_data(as_text=True)
    check("F 手機作廢單明細顯示已作廢、無狀態表單", "已作廢" in mdet and 'name="shipping_status"' not in mdet)

    # 手機出貨流程 → 已出貨單：staff 403 / owner 可
    sc.post("/m/orders/new", data={"product_id": str(pid), "combo_code": "BOX", "qty": "1",
                                   "amount": "300"}, follow_redirects=True)
    oid4, ono4 = last_order_id()
    box_o4 = bal(pid, "boxed")
    sc.post(f"/m/shipments/{oid4}/ship", data={}, follow_redirects=True)
    check("F 手機出貨建 shipments 列", shipments_count(oid4) == 1)
    r = sc.post(f"/m/orders/{oid4}/void", data={"reason": "x"})
    check("F 手機已出貨單 staff 作廢 403", r.status_code == 403 and order(oid4).voided_at is None)
    r = oc.post(f"/m/orders/{oid4}/void", data={"reason": "owner"}, follow_redirects=True)
    check("F 手機已出貨單 owner 作廢不回補", order(oid4).voided_at is not None and bal(pid, "boxed") == box_o4)

    # =====================================================================
    # H. 已售統計 = SALE − 回補淨額；viewer 不能作廢
    # =====================================================================
    from blueprints.reports import _inventory_report
    db = fresh()
    rows = _inventory_report(db)
    sold = sum(r["sold"] for r in rows if r["product_id"] == pid)
    db.close()
    expect_sold = (box0 - bal(pid, "boxed")) + (loose0 - bal(pid, "loose_piece"))
    check("H 庫存報表已售 = 淨銷售（SALE−SALE_REVERSAL）", sold == expect_sold,
          f"sold={sold} expect={expect_sold}")
    check("H audit_logs 總數：order_edit 3 / order_void 4", audit_count("order_edit") == 3
          and audit_count("order_void") == 4, f"edit={audit_count('order_edit')} void={audit_count('order_void')}")
    vc = app.test_client()
    login(vc, "viewer", "viewer123")
    sc.post("/orders/new", data={"item_product_id": str(pid), "item_combo_code": "BOX",
                                 "item_qty": "1", "item_amount": "300"}, follow_redirects=True)
    oid5, _ = last_order_id()
    r = vc.post(f"/orders/{oid5}/void", data={"reason": "viewer"})
    check("H viewer 作廢 403", r.status_code == 403)
    r = vc.get(f"/orders/{oid5}/edit")
    check("H viewer 編輯 403", r.status_code == 403)
    det5 = oc.get(f"/orders/{oid5}").get_data(as_text=True)
    check("H 桌面明細有編輯與作廢入口", f"/orders/{oid5}/edit" in det5 and f"/orders/{oid5}/void" in det5)

    print("\n==== 結果 ====")
    print(f"PASS={len(PASS)}  FAIL={len(FAIL)}")
    if FAIL:
        print("FAILED:")
        for f in FAIL:
            print("  - " + f)
        print("\n總結：FAIL")
        return 1
    print("\n總結：PASS")
    return 0


if __name__ == "__main__":
    sys.exit(main())
