# -*- coding: utf-8 -*-
"""CR-8 驗收：操作紀錄（audit_logs 補齊寫入 + /admin/audit 查閱頁 + Excel 分頁）— 2026-08-31。

throwaway sqlite 放 %TEMP%，不污染交付庫；env 在 import 任何 app 模組前先設定。
驗證項目：
  A 事件寫入：建單(order_create) / 改付款 / 改出貨(桌面+手機) / 設定變更 / 客戶新增+編輯+刪除 /
    商品編輯 / 支出新增 / 庫存補貨 / 登入成功+失敗 → 各產生對應 audit，detail 含 before/after
  B 權限：/admin/audit owner 200、accounting 200、staff 403、viewer 403、未登入導向登入
  C 篩選：帳號（只剩該帳號）、動作類型、日期 from/to（未來日期 0 筆、今日含全部）、關鍵字（單號）
  D 顯示：action 中文（建立訂單／改付款狀態…）、摘要含 before → after、對象連到訂單明細、<details> 原始 JSON、分頁
  E 手機 /m/audit：owner 200、staff 403
  F Excel：export_all 含「操作紀錄」分頁、表頭正確、含建單列
  G 密碼類不寫值：login 的 detail 不含 password 字串
"""
import os
import sys
import json
import tempfile
from datetime import datetime, timedelta, timezone
from io import BytesIO

try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except Exception:
    pass

_TMP_DB = os.path.join(tempfile.gettempdir(), "flora_court_audit_test.db")
if os.path.exists(_TMP_DB):
    os.remove(_TMP_DB)
os.environ["DATABASE_URL"] = "sqlite:///" + _TMP_DB.replace("\\", "/")
os.environ["ENABLE_MOBILE"] = "true"
os.environ["SECRET_KEY"] = "test-secret"

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import init_db  # noqa: E402
from db import get_session, Order, Product, Customer, Setting, AuditLog, User  # noqa: E402
from app import create_app  # noqa: E402

PASS = []
FAIL = []


def check(name, cond, detail=""):
    (PASS if cond else FAIL).append(name)
    print(f"[{'PASS' if cond else 'FAIL'}] {name}" + (f" — {detail}" if detail else ""))


def login(client, username, password):
    return client.post("/login", data={"username": username, "password": password},
                       follow_redirects=True)


def fresh():
    s = get_session()
    s.expire_all()
    return s


def audits(action, target_id=None):
    db = fresh()
    q = db.query(AuditLog).filter_by(action=action)
    if target_id is not None:
        q = q.filter_by(target_id=str(target_id))
    return q.order_by(AuditLog.id.asc()).all()


def det(a):
    return json.loads(a.detail or "{}")


def main():
    print(f"=== throwaway DB: {_TMP_DB} ===")
    init_db.create_all()
    db0 = get_session()
    init_db.seed(db0)
    db0.close()

    app = create_app()
    app.config["TESTING"] = True
    owner = app.test_client()
    staff = app.test_client()
    acct = app.test_client()
    viewer = app.test_client()
    anon = app.test_client()

    # ---- G/A 登入留痕 ----
    r = owner.post("/login", data={"username": "owner", "password": "wrong-pw"}, follow_redirects=True)
    fails = audits("login_fail")
    check("A 登入失敗 → login_fail", len(fails) == 1 and fails[0].actor_name == "owner"
          and fails[0].actor_id is None, f"n={len(fails)}")
    check("G login_fail detail 不含密碼", fails and "wrong-pw" not in (fails[0].detail or "")
          and "password" not in (fails[0].detail or ""))
    login(owner, "owner", "owner123")
    login(staff, "staff", "staff123")
    login(acct, "accounting", "accounting123")
    login(viewer, "viewer", "viewer123")
    oks = audits("login_ok")
    check("A 登入成功 → login_ok（4 帳號）", len(oks) == 4 and all(a.actor_id for a in oks), f"n={len(oks)}")
    check("G login_ok detail 不含密碼", all("123" not in (a.detail or "") for a in oks))

    db = fresh()
    mask = db.query(Product).filter_by(is_packaging=False).order_by(Product.id).first()
    owner_user = db.query(User).filter_by(username="owner").first()
    staff_user = db.query(User).filter_by(username="staff").first()
    pid = mask.id
    mask_name = mask.name
    owner_uid = owner_user.id
    staff_uid = staff_user.id

    # ---- A1 建單（staff 桌面）----
    r = staff.post("/orders/new", data={
        "recipient_name": "王小明", "new_customer_name": "稽核客戶甲",
        "item_product_id": [str(pid)], "item_combo_code": ["BOX"],
        "item_qty": ["1"], "item_amount": ["600"], "discount": "50",
    }, follow_redirects=True)
    check("A1 桌面建單 200", r.status_code == 200 and "已建立" in r.get_data(as_text=True))
    db = fresh()
    o1 = db.query(Order).order_by(Order.id.desc()).first()
    o1_id, o1_no = o1.id, o1.order_no
    a = audits("order_create", o1_id)
    check("A1 order_create 寫入 1 筆", len(a) == 1, f"n={len(a)}")
    d = det(a[0]) if a else {}
    check("A1 order_create actor=staff", a and a[0].actor_id == staff_uid and a[0].actor_name == "員工",
          f"{a[0].actor_id if a else None}/{a[0].actor_name if a else None}")
    check("A1 order_create detail 含單號/客戶/品項/金額",
          d.get("order_no") == o1_no and d.get("customer_name") == "稽核客戶甲"
          and len(d.get("items", [])) == 1 and d["items"][0]["qty"] == 1
          and float(d.get("total_amount")) == 550.0 and d.get("created_customer") == "稽核客戶甲", str(d)[:200])

    # ---- A1b 手機建單 ----
    r = staff.post("/m/orders/new", data={
        "product_id": str(pid), "combo_code": "LOOSE", "qty": "2", "amount": "200",
    }, follow_redirects=True)
    check("A1b 手機建單 200", r.status_code == 200 and "已建立" in r.get_data(as_text=True))
    db = fresh()
    o2 = db.query(Order).order_by(Order.id.desc()).first()
    o2_id = o2.id
    a = audits("order_create", o2_id)
    d = det(a[0]) if a else {}
    check("A1b 手機 order_create 寫入且 via=mobile", len(a) == 1 and d.get("via") == "mobile"
          and d.get("items", [{}])[0].get("combo_code") == "LOOSE", str(d)[:160])

    # ---- A2 改付款（桌面）----
    r = staff.post(f"/orders/{o1_id}/payment", data={"payment_status": "paid"}, follow_redirects=True)
    a = audits("order_payment_status", o1_id)
    d = det(a[0]) if a else {}
    check("A2 order_payment_status 寫入 before/after", len(a) == 1
          and d.get("before") == {"payment_status": "unpaid"} and d.get("after") == {"payment_status": "paid"}, str(d))
    # 同值再送一次 → 不多寫
    staff.post(f"/orders/{o1_id}/payment", data={"payment_status": "paid"}, follow_redirects=True)
    check("A2 付款狀態未變 → 不重複留痕", len(audits("order_payment_status", o1_id)) == 1)

    # ---- A2b 改付款（手機）----
    r = staff.post(f"/m/orders/{o2_id}", data={"payment_status": "partial"}, follow_redirects=True)
    a = audits("order_payment_status", o2_id)
    d = det(a[0]) if a else {}
    check("A2b 手機改付款 → order_payment_status via=mobile", len(a) == 1 and d.get("via") == "mobile"
          and d.get("after") == {"payment_status": "partial"}, str(d))

    # ---- A3 改出貨（桌面 update_shipping）----
    r = staff.post(f"/orders/{o1_id}/shipping", data={"shipping_status": "shipped", "tracking_no": "TRK001"},
                   follow_redirects=True)
    a = audits("order_shipping_status", o1_id)
    d = det(a[0]) if a else {}
    check("A3 桌面出貨 → order_shipping_status before/after + 紙袋", len(a) == 1
          and d.get("before") == {"shipping_status": "pending"} and d.get("after") == {"shipping_status": "shipped"}
          and d.get("paper_bag_deducted") is True and d.get("tracking_no") == "TRK001", str(d))

    # ---- A3b 手機出貨 ship_order ----
    r = staff.post(f"/m/shipments/{o2_id}/ship", data={"tracking_no": "TRK002", "carrier": "黑貓"},
                   follow_redirects=True)
    a = audits("order_shipping_status", o2_id)
    d = det(a[0]) if a else {}
    check("A3b 手機出貨 → order_shipping_status via=mobile", len(a) == 1 and d.get("via") == "mobile"
          and d.get("after") == {"shipping_status": "shipped"} and d.get("carrier") == "黑貓", str(d))

    # ---- A4 設定變更（owner）----
    db = fresh()
    s = db.query(Setting).filter_by(key="paperbag_qty_per_shipment").first()
    old_val = s.value
    r = owner.post("/products/settings/update", data={"key": "paperbag_qty_per_shipment", "value": "2"},
                   follow_redirects=True)
    a = audits("setting_update", "paperbag_qty_per_shipment")
    d = det(a[0]) if a else {}
    check("A4 setting_update before/after", len(a) == 1 and d.get("before") == {"value": old_val}
          and d.get("after") == {"value": "2"} and a[0].actor_id == owner_uid, str(d))
    owner.post("/products/settings/update", data={"key": "paperbag_qty_per_shipment", "value": "2"},
               follow_redirects=True)
    check("A4 設定值未變 → 不重複留痕", len(audits("setting_update", "paperbag_qty_per_shipment")) == 1)
    # 還原
    owner.post("/products/settings/update", data={"key": "paperbag_qty_per_shipment", "value": old_val},
               follow_redirects=True)

    # ---- A5 客戶新增／編輯／刪除（staff）----
    r = staff.post("/customers/new", data={"name": "稽核客戶乙", "phone": "0911222333"}, follow_redirects=True)
    db = fresh()
    c = db.query(Customer).filter_by(name="稽核客戶乙").first()
    c_id = c.id if c else -1
    a = audits("customer_create", c_id)
    check("A5 customer_create 寫入（after 含 name/phone）", c is not None and len(a) == 1
          and det(a[0]).get("after", {}).get("phone") == "0911222333")
    r = staff.post(f"/customers/{c_id}/edit", data={"name": "稽核客戶乙", "phone": "0999888777", "note": "VIP"},
                   follow_redirects=True)
    a = audits("customer_update", c_id)
    d = det(a[0]) if a else {}
    check("A5 customer_update 只寫有變欄位（phone/note，不含 name）", len(a) == 1
          and d.get("before", {}).get("phone") == "0911222333" and d.get("after", {}).get("phone") == "0999888777"
          and d.get("after", {}).get("note") == "VIP" and "name" not in d.get("after", {}), str(d))
    r = staff.post(f"/customers/{c_id}/delete", follow_redirects=True)
    a = audits("customer_delete", c_id)
    check("A5 customer_delete 寫入（before 快照）", len(a) == 1 and det(a[0]).get("before", {}).get("name") == "稽核客戶乙")

    # ---- A6 商品編輯（owner）----
    r = owner.post(f"/products/{pid}/edit", data={"name": mask_name, "category": "稽核類別", "active": "on"},
                   follow_redirects=True)
    a = audits("product_update", pid)
    d = det(a[-1]) if a else {}
    check("A6 product_update before/after（category）", len(a) >= 1 and d.get("after", {}).get("category") == "稽核類別"
          and "category" in d.get("before", {}), str(d))

    # ---- A7 支出新增（accounting）----
    r = acct.post("/reports/expenses/add", data={"name": "稽核支出", "category": "其他", "amount": "123",
                                                  "expense_date": "2026-08-31"}, follow_redirects=True)
    a = audits("expense_create")
    d = det(a[-1]) if a else {}
    check("A7 expense_create（accounting）", len(a) == 1 and d.get("after", {}).get("name") == "稽核支出"
          and a[0].actor_name == "會計", str(d)[:120])

    # ---- A8 庫存補貨（staff）----
    r = staff.post("/inventory/restock", data={"product_id": str(pid), "inventory_pool": "boxed",
                                               "stock_category": "normal", "movement_type": "PURCHASE",
                                               "qty": "3", "note": "稽核補貨"}, follow_redirects=True)
    a = audits("inventory_restock", pid)
    d = det(a[0]) if a else {}
    check("A8 inventory_restock 含 qty_before/qty_after/movement_ids", len(a) == 1
          and d.get("qty_after") == (d.get("qty_before") or 0) + 3 and d.get("movement_ids")
          and d.get("movement_type") == "PURCHASE", str(d)[:160])

    # ---- B 權限 ----
    r = owner.get("/admin/audit/")
    check("B owner /admin/audit 200", r.status_code == 200, f"got {r.status_code}")
    html_owner = r.get_data(as_text=True)
    r = acct.get("/admin/audit/")
    check("B accounting /admin/audit 200", r.status_code == 200, f"got {r.status_code}")
    r = staff.get("/admin/audit/")
    check("B staff /admin/audit 403", r.status_code == 403, f"got {r.status_code}")
    r = viewer.get("/admin/audit/")
    check("B viewer /admin/audit 403", r.status_code == 403, f"got {r.status_code}")
    r = anon.get("/admin/audit/")
    check("B 未登入導向 /login", r.status_code == 302 and "/login" in r.headers.get("Location", ""))
    check("B 導覽列 owner 見「操作紀錄」", "操作紀錄" in html_owner and 'href="/admin/audit/"' in html_owner)
    r = staff.get("/orders/")
    check("B 導覽列 staff 不見「操作紀錄」入口", 'href="/admin/audit/"' not in r.get_data(as_text=True))

    # ---- D 顯示 ----
    check("D action 中文：建立訂單", "建立訂單" in html_owner)
    check("D action 中文：改付款狀態", "改付款狀態" in html_owner)
    check("D action 中文：改出貨狀態", "改出貨狀態" in html_owner)
    check("D action 中文：變更設定／登入失敗", "變更設定" in html_owner and "登入失敗" in html_owner)
    check("D 摘要含 before → after（未付款 → 已付款）", "未付款 → 已付款" in html_owner)
    check("D 摘要含出貨 待出貨 → 已出貨", "待出貨 → 已出貨" in html_owner)
    check("D 對象連到訂單明細", f'href="/orders/{o1_id}"' in html_owner)
    check("D <details> 原始 JSON", "<details>" in html_owner and '&#34;before&#34;' in html_owner or '"before"' in html_owner)
    check("D 顯示帳號（員工）", "員工" in html_owner)
    check("D 時間降冪（首列不是 login_fail）",
          html_owner.find("登入失敗") > html_owner.find("建立訂單"))

    # ---- C 篩選 ----
    r = owner.get(f"/admin/audit/?actor={staff_uid}")
    h = r.get_data(as_text=True)
    check("C 帳號篩選 staff：有建立訂單、無變更設定（表格列）",
          '<td class="act">建立訂單' in h and '<td class="act">變更設定' not in h)
    r = owner.get("/admin/audit/?action=order_payment_status")
    h = r.get_data(as_text=True)
    check("C 動作篩選：只剩改付款狀態", "改付款狀態" in h and "<td class=\"act\">建立訂單" not in h)
    # 日期：台北今日含全部；未來日期 0 筆
    today_tpe = (datetime.now(timezone.utc) + timedelta(hours=8)).strftime("%Y-%m-%d")
    future = (datetime.now(timezone.utc) + timedelta(days=3)).strftime("%Y-%m-%d")
    r = owner.get(f"/admin/audit/?from={today_tpe}&to={today_tpe}")
    h = r.get_data(as_text=True)
    check("C 日期篩選（今日）含建立訂單", "建立訂單" in h)
    r = owner.get(f"/admin/audit/?from={future}")
    h = r.get_data(as_text=True)
    check("C 日期篩選（未來）0 筆", "沒有符合條件的紀錄" in h and "共 0 筆" in h)
    r = owner.get(f"/admin/audit/?q={o1_no}")
    h = r.get_data(as_text=True)
    check("C 關鍵字（單號）命中建單與狀態變更、不含登入列", o1_no in h
          and '<td class="act">建立訂單' in h and '<td class="act">改付款狀態' in h
          and '<td class="act">登入失敗' not in h)
    r = owner.get("/admin/audit/?from=bad-date")
    check("C 日期格式錯 → 200 + 提示", r.status_code == 200 and "日期格式" in r.get_data(as_text=True))

    # ---- D 分頁：塞 60 筆後第 2 頁存在 ----
    from audit_util import write_audit
    db = fresh()
    for i in range(60):
        write_audit(db, "product_update", "products", pid, {"before": {"note": i}, "after": {"note": i + 1}},
                    actor_id=owner_uid, actor_name="老闆")
    db.commit()
    r = owner.get("/admin/audit/")
    h = r.get_data(as_text=True)
    total_now = fresh().query(AuditLog).count()
    check("D 分頁：每頁 50、顯示總數", f"共 {total_now} 筆" in h and "第 1 /" in h and "下一頁" in h)
    r = owner.get("/admin/audit/?page=2")
    check("D 分頁：第 2 頁 200", r.status_code == 200 and "第 2 /" in r.get_data(as_text=True))

    # ---- E 手機 ----
    r = owner.get("/m/audit")
    h = r.get_data(as_text=True)
    check("E /m/audit owner 200 且含建立訂單", r.status_code == 200 and "建立訂單" in h)
    r = staff.get("/m/audit")
    check("E /m/audit staff 403", r.status_code == 403, f"got {r.status_code}")
    r = owner.get("/m/")
    check("E 手機首頁 owner 見操作紀錄入口", 'href="/m/audit"' in r.get_data(as_text=True))
    r = staff.get("/m/")
    check("E 手機首頁 staff 不見操作紀錄入口", 'href="/m/audit"' not in r.get_data(as_text=True))

    # ---- F Excel ----
    r = owner.get("/reports/export.xlsx")
    check("F 匯出 200", r.status_code == 200)
    from openpyxl import load_workbook
    wb = load_workbook(BytesIO(r.get_data()))
    check("F 含「操作紀錄」分頁", "操作紀錄" in wb.sheetnames, str(wb.sheetnames))
    if "操作紀錄" in wb.sheetnames:
        ws = wb["操作紀錄"]
        header = [c.value for c in ws[1]]
        check("F 表頭", header == ["時間", "帳號", "動作", "對象類型", "對象ID", "摘要", "詳細(JSON)"], str(header))
        flat = [str(c.value) for row in ws.iter_rows(min_row=2) for c in row if c.value is not None]
        check("F 含建立訂單列與單號", any("建立訂單" in v for v in flat) and any(o1_no in v for v in flat))
        check("F 列數 ≤ 1000 + 表頭", ws.max_row <= 1001)

    print("\n=== 結果 ===")
    print(f"PASS: {len(PASS)}  FAIL: {len(FAIL)}")
    if FAIL:
        print("FAILED:")
        for f in FAIL:
            print("  - " + f)
        sys.exit(1)
    print("ALL GREEN")


if __name__ == "__main__":
    main()
