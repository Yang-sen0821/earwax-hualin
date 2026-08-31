# -*- coding: utf-8 -*-
"""CR-5 驗收：訂單運費 / 運送方式 / 運送備註 / 折扣落地（2026-08-31）。

throwaway sqlite 放 %TEMP%，不污染交付庫；env 在 import 任何 app 模組前先設定。
驗證項目：
  ① 桌面建單帶運費/方式/備註/折扣 → 欄位落地；total_amount = 商品實收 − 折扣（不含運費）
  ② 手機建單帶運費/方式 → 欄位落地；total 不含運費
  ③ 不帶新欄 → 預設 0 / 空（舊 POST 相容）
  ④ 運送方式非閉集 → 擋下、不建單
  ⑤ 明細（桌面/手機）、列表、期間報表、Excel 顯示運費 / 方式 / 應收合計
"""
import os
import sys
import io
import tempfile

try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except Exception:
    pass

_TMP_DB = os.path.join(tempfile.gettempdir(), "flora_court_shipfee_test.db")
if os.path.exists(_TMP_DB):
    os.remove(_TMP_DB)
os.environ["DATABASE_URL"] = "sqlite:///" + _TMP_DB.replace("\\", "/")
os.environ["ENABLE_MOBILE"] = "true"
os.environ["SECRET_KEY"] = "test-secret"

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import init_db  # noqa: E402
from db import get_session, Order, Product  # noqa: E402
from app import create_app  # noqa: E402

PASS = []
FAIL = []


def check(name, cond, detail=""):
    (PASS if cond else FAIL).append(name)
    print(f"[{'PASS' if cond else 'FAIL'}] {name}" + (f" — {detail}" if detail else ""))


def login(client, username="owner", password="owner123"):
    return client.post("/login", data={"username": username, "password": password},
                       follow_redirects=True)


def last_order():
    db = get_session()
    db.expire_all()
    o = db.query(Order).order_by(Order.id.desc()).first()
    db.close()
    return o


def main():
    init_db.create_all()
    s = get_session()
    try:
        init_db.seed(s)
    finally:
        s.close()

    app = create_app()
    app.config["TESTING"] = True
    client = app.test_client()
    login(client)

    db = get_session()
    pid = db.query(Product).filter_by(sku="FC-MASK-001").first().id
    n0 = db.query(Order).count()
    db.close()

    # ---- ① 桌面建單帶運費/方式/備註/折扣 ----
    r = client.post("/orders/new", data={
        "recipient_name": "王小明",
        "item_product_id": [str(pid), str(pid)],
        "item_combo_code": ["BOX", "LOOSE"],
        "item_qty": ["1", "2"],
        "item_amount": ["600", "300"],
        "discount": "50",
        "shipping_fee": "60",
        "shipping_method": "711",
        "shipping_note": "7-11 測試門市",
    }, follow_redirects=True)
    check("①桌面建單成功(200)", r.status_code == 200 and "已建立" in r.get_data(as_text=True))
    o = last_order()
    check("①shipping_fee=60 落地", float(o.shipping_fee) == 60.0, f"got={o.shipping_fee}")
    check("①shipping_method=711 落地", o.shipping_method == "711", f"got={o.shipping_method}")
    check("①shipping_note 落地", o.shipping_note == "7-11 測試門市", f"got={o.shipping_note}")
    check("①discount=50 落地", float(o.discount) == 50.0, f"got={o.discount}")
    check("①total_amount = 900−50 = 850（不含運費）", float(o.total_amount) == 850.0,
          f"got={o.total_amount}")
    o1_id = o.id

    det = client.get(f"/orders/{o1_id}").get_data(as_text=True)
    check("①明細顯示運送方式中文『711 店到店』", "711 店到店" in det)
    check("①明細顯示運費 60", "運費" in det and "60" in det)
    check("①明細顯示應收合計 910", "應收合計" in det and "910" in det)
    check("①明細顯示運送備註門市名", "7-11 測試門市" in det)

    # ---- ② 手機建單帶運費/方式 ----
    sc = app.test_client()
    login(sc, "staff", "staff123")
    r = sc.post("/m/orders/new", data={
        "product_id": str(pid), "combo_code": "BOX", "qty": "1", "amount": "300",
        "shipping_fee": "80", "shipping_method": "post", "shipping_note": "掛號",
    }, follow_redirects=True)
    check("②手機建單成功", r.status_code == 200 and "已建立" in r.get_data(as_text=True))
    o = last_order()
    check("②手機 shipping_fee=80", float(o.shipping_fee) == 80.0, f"got={o.shipping_fee}")
    check("②手機 shipping_method=post", o.shipping_method == "post")
    check("②手機 shipping_note=掛號", o.shipping_note == "掛號")
    check("②手機 total=300（不含運費）", float(o.total_amount) == 300.0, f"got={o.total_amount}")
    check("②手機 discount 預設 0", float(o.discount or 0) == 0.0)
    o2_id = o.id
    mdet = sc.get(f"/m/orders/{o2_id}").get_data(as_text=True)
    check("②手機明細顯示『郵寄』與運費 80", "郵寄" in mdet and "$80" in mdet)
    check("②手機明細顯示應收合計 380", "應收合計" in mdet and "$380" in mdet)

    # ---- ③ 不帶新欄 → 預設 ----
    r = client.post("/orders/new", data={
        "item_product_id": str(pid), "item_combo_code": "BOX",
        "item_qty": "1", "item_amount": "300",
    }, follow_redirects=True)
    o = last_order()
    check("③桌面不帶新欄 → shipping_fee=0", float(o.shipping_fee or 0) == 0.0, f"got={o.shipping_fee}")
    check("③桌面不帶新欄 → shipping_method 空", o.shipping_method is None)
    check("③桌面不帶新欄 → discount=0 / total=300", float(o.discount or 0) == 0.0
          and float(o.total_amount) == 300.0)
    r = sc.post("/m/orders/new", data={
        "product_id": str(pid), "combo_code": "LOOSE", "qty": "1", "amount": "50",
    }, follow_redirects=True)
    o = last_order()
    check("③手機不帶新欄 → 預設 0 / 空", float(o.shipping_fee or 0) == 0.0
          and o.shipping_method is None and o.shipping_note is None)

    # ---- ④ 非閉集運送方式 → 擋下 ----
    db = get_session()
    n_before = db.query(Order).count()
    db.close()
    r = client.post("/orders/new", data={
        "item_product_id": str(pid), "item_combo_code": "BOX",
        "item_qty": "1", "item_amount": "300", "shipping_method": "ufo",
    }, follow_redirects=True)
    db = get_session()
    check("④桌面非閉集運送方式 → 不建單且提示", db.query(Order).count() == n_before
          and "無效的運送方式" in r.get_data(as_text=True))
    db.close()
    r = sc.post("/m/orders/new", data={
        "product_id": str(pid), "combo_code": "BOX", "qty": "1", "amount": "300",
        "shipping_method": "ufo",
    }, follow_redirects=True)
    db = get_session()
    check("④手機非閉集運送方式 → 不建單且提示", db.query(Order).count() == n_before
          and "無效的運送方式" in r.get_data(as_text=True))
    db.close()

    # ---- ⑤ 列表 / 報表 / Excel ----
    lst = client.get("/orders/").get_data(as_text=True)
    check("⑤桌面列表有運費/運送方式欄", "運費" in lst and "運送方式" in lst and "711 店到店" in lst)
    mlst = sc.get("/m/orders").get_data(as_text=True)
    check("⑤手機列表顯示運費與方式", "運費 $80" in mlst and "郵寄" in mlst)
    form_html = client.get("/orders/new").get_data(as_text=True)
    check("⑤桌面建單頁有三欄", 'name="shipping_fee"' in form_html
          and 'name="shipping_method"' in form_html and 'name="shipping_note"' in form_html)
    mform = sc.get("/m/orders/new").get_data(as_text=True)
    check("⑤手機建單頁有三欄", 'name="shipping_fee"' in mform
          and 'name="shipping_method"' in mform and 'name="shipping_note"' in mform)
    check("⑤建單下拉含 4 種運送方式中文", all(x in form_html for x in
          ("711 店到店", "郵寄", "自取", "其他")))

    sales = client.get("/reports/sales?g=month").get_data(as_text=True)
    check("⑤期間報表有運費/應收合計欄", "運費" in sales and "應收合計" in sales)
    # 期間運費合計 = 60 + 80 = 140
    check("⑤期間報表運費合計 140", "$140" in sales)

    xlsx = client.get("/reports/export.xlsx")
    check("⑤Excel 200", xlsx.status_code == 200)
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(xlsx.data))
    ws = wb["訂單"]
    headers = [c.value for c in ws[1]]
    check("⑤Excel 訂單分頁表頭含 運費/運送方式/應收合計/折扣",
          all(h in headers for h in ("運費", "運送方式", "應收合計", "折扣")), str(headers))
    rows = list(ws.iter_rows(values_only=True))[1:]
    hi = {h: i for i, h in enumerate(headers)}
    row1 = [r for r in rows if r[hi["運費"]] == 60.0]
    check("⑤Excel 首單運費 60 / 應收合計 910 / 方式 711 店到店",
          bool(row1) and row1[0][hi["應收合計"]] == 910.0 and row1[0][hi["運送方式"]] == "711 店到店"
          and row1[0][hi["總金額"]] == 850.0 and row1[0][hi["折扣"]] == 50.0,
          str(row1[0]) if row1 else "no row")
    ws_s = wb["銷售報表"]
    flat = [str(c.value) for row in ws_s.iter_rows() for c in row if c.value is not None]
    check("⑤Excel 銷售分頁期間表含 運費 欄", "運費" in flat and "應收合計" in flat)

    print("\n==== 結果 ====")
    print(f"PASS={len(PASS)}  FAIL={len(FAIL)}")
    if FAIL:
        print("FAILED:")
        for f in FAIL:
            print("  - " + f)
        print("\n總結：FAIL")
        return 1
    print("\n總結：PASS")
    return 0


if __name__ == "__main__":
    sys.exit(main())
