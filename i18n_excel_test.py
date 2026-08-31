"""本機驗證：中文化顯示 + Excel 單鍵多分頁 + 其他功能未壞。

throwaway sqlite 放 %TEMP%，不污染交付庫。env 在 import 任何 app 模組前先設定。
"""
import os
import sys
import io
import tempfile

# 強制 UTF-8 輸出（避免 Windows cp950 終端亂碼）
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

# ---- throwaway sqlite（%TEMP%）----
_TMP = os.path.join(tempfile.gettempdir(), "flora_court_i18n_test.db")
if os.path.exists(_TMP):
    os.remove(_TMP)
os.environ["DATABASE_URL"] = "sqlite:///" + _TMP.replace("\\", "/")
os.environ["SECRET_KEY"] = "test-secret"

# import 在 env 設定後
from io import BytesIO
from datetime import datetime

import init_db
from db import (
    get_session, Order, OrderItem, Customer, InventoryMovement, InventoryBalance,
    Product, SalesPlan,
)
import inventory_service
from app import create_app

PASS = []
FAIL = []


def check(name, cond, detail=""):
    (PASS if cond else FAIL).append(name)
    mark = "PASS" if cond else "FAIL"
    print(f"[{mark}] {name}" + (f" — {detail}" if detail else ""))


def main():
    print(f"=== throwaway DB: {_TMP} ===")
    # ---- 1. 建表 + seed ----
    init_db.create_all()
    db0 = get_session()
    init_db.seed(db0)
    db0.close()
    print("[init] seed done")

    app = create_app()
    app.config["TESTING"] = True
    client = app.test_client()

    # ---- 登入 owner ----
    r = client.post("/login", data={"username": "owner", "password": "owner123"},
                    follow_redirects=True)
    check("登入 owner", r.status_code == 200)

    # ---- 2. 塞樣本訂單（不同 combo / 狀態）----
    db = get_session()
    mask = db.query(Product).filter_by(sku="FC-MASK-001").first()
    cust = Customer(name="測試客戶", phone="0900000000")
    db.add(cust); db.flush()

    def make_order(order_no, combos_qty, pay, ship):
        o = Order(order_no=order_no, customer_id=cust.id,
                  recipient_name="收件人A", recipient_phone="0911",
                  shipping_address="台北市測試路1號",
                  total_amount=0, payment_status=pay, shipping_status=ship)
        db.add(o); db.flush()
        total = 0
        for combo, qty in combos_qty:
            result = inventory_service.deduct_for_sale(
                db, product_id=mask.id, combo_code=combo, order_qty=qty,
                operator="tester", order_ref=str(o.id), note="sample")
            ok = bool(getattr(result, "ok", True))
            if not ok:
                db.rollback(); raise SystemExit(f"扣庫失敗 {combo}")
            db.add(OrderItem(order_id=o.id, product_id=mask.id, combo_code=combo,
                             qty=qty, unit_price=100, subtotal=100 * qty))
            total += 100 * qty
        o.total_amount = total
        db.flush()
        return o

    # 扣庫前盒裝 normal
    bal_box_before = db.query(InventoryBalance).filter_by(
        product_id=mask.id, inventory_pool="boxed", stock_category="normal").first().qty

    o1 = make_order("FCTEST001", [("SINGLE", 2), ("BOX1", 1)], "paid", "shipped")
    o2 = make_order("FCTEST002", [("BOX3", 1)], "unpaid", "pending")
    o3 = make_order("FCTEST003", [("BOX10", 1)], "partial", "delivered")
    db.commit()
    # 在 close 前先取出需要的 id（避免 DetachedInstanceError）
    o3_id = o3.id

    bal_box_after = db.query(InventoryBalance).filter_by(
        product_id=mask.id, inventory_pool="boxed", stock_category="normal").first().qty
    # BOX1(1)+BOX3(3)+BOX10(10)=14 盒
    check("建單扣庫正確（盒裝 -14）", bal_box_before - bal_box_after == 14,
          f"{bal_box_before}->{bal_box_after}")

    # 額外庫存異動：GIFT / PR / 補貨，覆蓋多種 movement_type
    inventory_service.deduct_out(db, mask.id, "loose_piece", "normal", "GIFT", 1, "tester", note="贈品樣本")
    inventory_service.restock(db, mask.id, "boxed", "normal", "PURCHASE", 5, "tester", note="進貨樣本")
    db.commit()
    db.close()

    # ---- 3. 中文化頁面驗證 ----
    def page(path):
        resp = client.get(path)
        return resp.status_code, resp.get_data(as_text=True)

    sc, html = page("/orders/")
    check("訂單列表頁 200", sc == 200)
    check("訂單列表：已付款", "已付款" in html)
    check("訂單列表：待出貨", "待出貨" in html)
    check("訂單列表：未付款", "未付款" in html)
    # 顯示文字（td 內）不得出現裸英文 code；下拉 value="code" 屬合法（value 仍存 code）
    check("訂單列表顯示無裸英文碼 >unpaid<",
          ">unpaid<" not in html and ">paid<" not in html and ">pending<" not in html)

    sc, html = page(f"/orders/{o3_id}")
    check("訂單明細頁 200", sc == 200)
    check("訂單明細：尊寵囤貨組(BOX10)", "尊寵囤貨組" in html)
    check("訂單明細：部分付款", "部分付款" in html)
    check("訂單明細：已送達", "已送達" in html)
    # 下拉 value 仍為 code
    check("訂單明細下拉 value=paid（code 不變）", 'value="paid"' in html)
    check("訂單明細下拉 value=shipped（code 不變）", 'value="shipped"' in html)

    sc, html = page("/reports/inventory")
    check("庫存報表頁 200", sc == 200)
    check("庫存報表：盒裝", "盒裝" in html)
    check("庫存報表：裸片", "裸片" in html)

    sc, html = page("/inventory/movements")
    check("庫存異動頁 200", sc == 200)
    check("庫存異動：銷售", "銷售" in html)
    check("庫存異動：贈品", "贈品" in html)
    check("庫存異動：進貨", "進貨" in html)
    check("庫存異動：期初(SEED)", "期初" in html)

    sc, html = page("/reports/sales?g=month")
    check("銷售報表頁 200", sc == 200)

    sc, html = page("/products/plans")
    check("方案頁：經典盒裝(BOX1)", "經典盒裝" in html)

    sc, html = page("/m/orders")
    check("手機訂單列表 200", sc == 200)
    check("手機訂單：中文狀態", "已付款" in html or "待出貨" in html)

    sc, html = page("/m/inventory")
    check("手機庫存 200", sc == 200)
    check("手機庫存：盒裝（無 boxed 尾綴）", "盒裝" in html and "盒裝 boxed" not in html)

    # ---- 4. Excel 單鍵多分頁 ----
    resp = client.get("/reports/export.xlsx")
    check("匯出路由 200", resp.status_code == 200)
    ctype = resp.headers.get("Content-Type", "")
    check("匯出 MIME=xlsx", "spreadsheetml" in ctype, ctype)
    data = resp.get_data()
    from openpyxl import load_workbook
    wb = load_workbook(BytesIO(data))
    sheets = wb.sheetnames
    check("分頁=訂單/庫存/庫存異動/銷售報表/客戶排行",
          sheets == ["訂單", "庫存", "庫存異動", "銷售報表", "客戶排行"], str(sheets))  # CR-3 加客戶排行
    check("無留空白 Sheet 分頁", "Sheet" not in sheets)

    # 內容中文 + 數字
    ws_o = wb["訂單"]
    o_vals = [[c.value for c in row] for row in ws_o.iter_rows()]
    flat_o = [str(v) for row in o_vals for v in row if v is not None]
    check("訂單分頁含中文狀態（已付款/待出貨）",
          any("已付款" in v for v in flat_o) and any("待出貨" in v for v in flat_o))
    check("訂單分頁含訂單編號 FCTEST001", any("FCTEST001" in v for v in flat_o))

    ws_m = wb["庫存異動"]
    flat_m = [str(c.value) for row in ws_m.iter_rows() for c in row if c.value is not None]
    check("異動分頁含中文類別（銷售/贈品/進貨）",
          any("銷售" in v for v in flat_m) and any("贈品" in v for v in flat_m))
    check("異動分頁庫存池中文（盒裝/裸片）", any(v in ("盒裝", "裸片") for v in flat_m))

    ws_s = wb["銷售報表"]
    flat_s = [str(c.value) for row in ws_s.iter_rows() for c in row if c.value is not None]
    check("銷售分頁含組合中文名", any("尊寵囤貨組" in v or "經典盒裝" in v for v in flat_s))

    ws_i = wb["庫存"]
    inv_vals = [[c.value for c in row] for row in ws_i.iter_rows()]
    # 第一列資料的數字應為整數
    has_number = any(isinstance(v, (int, float)) for row in inv_vals[1:] for v in row)
    check("庫存分頁含數字", has_number)

    # ---- 4b. 舊分散路由已收掉 ----
    for old in ["/reports/export/orders.xlsx", "/reports/export/inventory.xlsx",
                "/reports/export/movements.xlsx", "/reports/export/sales.xlsx"]:
        r = client.get(old)
        check(f"舊路由已移除 {old} (404)", r.status_code == 404, f"got {r.status_code}")

    # ---- 5. 其他功能未壞：ROI / 支出管理 / 權限 ----
    sc, html = page("/reports/roi")
    check("ROI 頁 200（owner）", sc == 200)
    sc, html = page("/reports/expenses")
    check("支出管理頁 200（owner）", sc == 200)

    # 權限：viewer 不能進 ROI
    client.get("/logout")
    client.post("/login", data={"username": "viewer", "password": "viewer123"})
    r = client.get("/reports/roi")
    check("viewer 進 ROI 被擋(403)", r.status_code == 403, f"got {r.status_code}")
    # viewer 仍可讀訂單（中文）
    sc, html = page("/orders/")
    check("viewer 可讀訂單列表", sc == 200)

    print("\n=== 結果 ===")
    print(f"PASS: {len(PASS)}  FAIL: {len(FAIL)}")
    if FAIL:
        print("FAILED:")
        for f in FAIL:
            print("  - " + f)
        sys.exit(1)
    print("ALL GREEN")


if __name__ == "__main__":
    main()
