"""orders blueprint（/orders）— 訂單管理（D4）。

職責：
- 建立訂單（多品項，§4.2 同一原子 transaction）：建單 + 建明細 + 逐項呼叫
  inventory_service.deduct_for_sale 扣庫存 + 寫 movement；任一品項缺貨 ⇒ 整張 rollback。
- 查詢 / 明細。
- 改付款狀態 / 出貨狀態；出貨成立同一 transaction 內呼叫 deduct_for_shipment 扣紙袋（§4.4），
  紙袋不足 ⇒ rollback、出貨不成立。
- 含收件地址快照、訂單編號、金額 / 折扣。
- 歷史訂單匯入欄位頁（欄位齊，資料後補，D14）。
- CR-4（2026-08-31）：訂單編輯（同一 tx：reverse_sale 反沖 → 刪重建 items → 逐列 deduct_for_sale
  → 重算 total）與作廢（軟刪除 voided_at/by/reason；未出貨全額回補、已出貨不回補限 owner）；
  兩者皆寫 audit_logs。
- CR-10（2026-09-01）：單筆訂單編輯可「新增一行／刪除一行／改數量金額」（桌面 form.html 動態列 +
  手機簡版 /m/orders/<id>/edit）；核心抽成 perform_edit() 供桌面／手機共用；audit order_edit detail
  另附 lines_added / lines_removed / lines_changed 摘要。
  **「已出貨」判定改為：只有 shipments 表有該單列（真正走過出貨流程、可能扣過紙袋）才鎖品項／不回補；
  shipping_status 為 shipped/delivered 但無 shipments 列 → 視為未出貨**（線上 51 筆歷史單全由客戶以
  paid/delivered 直接建立、從未走出貨流程，舊規則把它們全鎖死，森哥要的加行／刪行做不到）。
- CR-7（2026-08-31）：GET /orders/new?customer_id=<id> 預填客戶（客戶明細頁「＋ 新增訂單」入口）；
  id 不存在則忽略、照常開空白建單頁；仍可換人（沿用客戶欄打字過濾）。
- CR-6（2026-08-31）：桌面列表勾選批次作廢（POST /orders/void-bulk）；規則完全沿用 perform_void，
  一個原因套用全批、整批同一 tx（任一失敗全數 rollback）、已作廢單跳過、staff 勾到已出貨單整批拒絕；
  每筆各寫一筆 AuditLog(order_void)，detail 另註 bulk=True / bulk_group_id（同批同一 uuid）。

紅線 R1（本模組對外契約遵守）：
- 任何改庫存量一律呼叫 inventory_service 的 §4 契約函式（deduct_for_sale / deduct_for_shipment），
  session 為首參、在同一 transaction 內操作；本模組「不」自己寫 inventory_balances /
  inventory_movements，「不」直接 UPDATE 庫存表。
- 訂單層所有扣減 movement 共用 ref_type=order + ref_id=order_id（由 service 內以 order_ref 標記）。

嚴守鐵律：只用本檔 + templates/orders/*；不改 db.py / app.py / base.html /
inventory_service.py / 其他 blueprint。
"""
import csv
import io
import json
import uuid
from datetime import datetime
from decimal import Decimal, InvalidOperation

from flask import (
    Blueprint, render_template, request, redirect, url_for, flash, abort,
)

from auth import login_required, role_required, current_user
from db import (
    get_session, COMBO_CODES,
    Order, OrderItem, Customer, CustomerAddress,
    Product, SalesPlan, Shipment,
    HistoricalOrderImport, HistoricalOrderImportRow,
    active_orders,
)
import inventory_service
from display_labels import combo_label
from audit_util import write_audit

orders_bp = Blueprint("orders", __name__, url_prefix="/orders")

# 可寫角色（owner 由 role_required 永遠通過；§六：orders staff=RW）
WRITE_ROLES = ("staff",)

# 狀態閉集（§二 orders 預設值對照；UI 下拉用）
PAYMENT_STATUSES = ("unpaid", "paid", "refunded", "partial")
# CR-4：移除 cancelled（原下拉可選取消但不回補庫存的漏洞），取消一律走「作廢」流程（void）
SHIPPING_STATUSES = ("pending", "shipped", "delivered")

# 建單品項單位閉集（新模型）：LOOSE=片(扣裸片池)、BOX=盒(扣盒裝池)、BAG=袋(扣紙袋池，2026-09-01)。
# combo_code 欄位沿用，存 LOOSE/BOX/BAG；數量=qty，金額=unit_price/subtotal（無 schema 變更）。
# 面膜商品只能選 片/盒；包材商品（is_packaging）只能選 袋 —— 由 validate_row_units 在後端擋。
UNIT_CODES = ("LOOSE", "BOX", "BAG")
MASK_UNIT_CODES = ("LOOSE", "BOX")
BAG_UNIT_CODES = ("BAG",)

# 運送方式閉集（CR-5 2026-08-31）；中文由 display_labels.shipping_method_label
SHIPPING_METHODS = ("711", "post", "pickup", "other")


# -------------------------------------------------------------------------
# 列表 / 查詢
# -------------------------------------------------------------------------
@orders_bp.route("/")
@login_required
def index():
    db = get_session()
    # 2026-09-01 客戶回報 bug：搜「陳筠涵」找不到 FC20260831041——該單收件人空白、
    # 名字在關聯客戶（customer_id）上，但搜尋只比對 order_no/recipient_*。
    # 修正：一併比對客戶姓名與客戶電話（outerjoin，未綁客戶的單不受影響）、
    # 去除前後半形/全形空白、改 ilike 忽略大小寫。
    q = (request.args.get("q") or "").strip().strip("　").strip()
    pay = (request.args.get("payment_status") or "").strip()
    ship = (request.args.get("shipping_status") or "").strip()
    show_voided = request.args.get("show_voided") == "1"

    query = db.query(Order)
    if not show_voided:
        query = active_orders(query)   # CR-4：預設不列作廢單；勾選後以灰字＋「已作廢」標示
    if q:
        like = f"%{q}%"
        query = query.outerjoin(Customer, Customer.id == Order.customer_id).filter(
            (Order.order_no.ilike(like))
            | (Order.recipient_name.ilike(like))
            | (Order.recipient_phone.ilike(like))
            | (Customer.name.ilike(like))
            | (Customer.phone.ilike(like))
        )
    if pay in PAYMENT_STATUSES:
        query = query.filter(Order.payment_status == pay)
    if ship in SHIPPING_STATUSES:
        query = query.filter(Order.shipping_status == ship)

    orders = query.order_by(Order.id.desc()).all()
    cust_map = {c.id: c for c in db.query(Customer).all()}
    u = current_user()
    # CR-6：owner / staff 可批次作廢（同單筆作廢的可寫角色；已出貨單限 owner 由路由層再擋）
    can_bulk_void = bool(u and (u.role == "owner" or u.role in WRITE_ROLES))
    return render_template(
        "orders/index.html", section="orders",
        orders=orders, cust_map=cust_map, q=q,
        payment_statuses=PAYMENT_STATUSES, shipping_statuses=SHIPPING_STATUSES,
        f_pay=pay, f_ship=ship, show_voided=show_voided,
        can_bulk_void=can_bulk_void,
    )


# -------------------------------------------------------------------------
# 明細
# -------------------------------------------------------------------------
@orders_bp.route("/<int:order_id>")
@login_required
def detail(order_id):
    db = get_session()
    order = db.get(Order, order_id)
    if not order:
        abort(404)
    items = db.query(OrderItem).filter_by(order_id=order_id).order_by(OrderItem.id.asc()).all()
    customer = db.get(Customer, order.customer_id) if order.customer_id else None
    prod_map = {p.id: p for p in db.query(Product).all()}
    u = current_user()
    role = u.role if u else ""
    shipped = is_order_shipped(db, order_id)
    return render_template(
        "orders/detail.html", section="orders",
        order=order, items=items, customer=customer, prod_map=prod_map,
        payment_statuses=PAYMENT_STATUSES, shipping_statuses=SHIPPING_STATUSES,
        is_shipped=shipped, is_voided=(order.voided_at is not None),
        can_write=(role in ("owner",) + WRITE_ROLES),
        can_void=(role == "owner" or (role in WRITE_ROLES and not shipped)),
    )


# -------------------------------------------------------------------------
# 建立訂單（多品項，§4.2 整張原子 transaction）
# -------------------------------------------------------------------------
@orders_bp.route("/new", methods=["GET", "POST"])
@role_required(*WRITE_ROLES)
def new():
    db = get_session()
    if request.method == "POST":
        return _create_order(db)
    return render_template(
        "orders/form.html", section="orders",
        customers=db.query(Customer).order_by(Customer.name.asc()).all(),
        products=sellable_products(db),
        unit_codes=UNIT_CODES, shipping_methods=SHIPPING_METHODS,
        form=_prefill_customer_form(db, request.args.get("customer_id")),
    )


def customer_display(cust):
    """客戶欄顯示字串（與 form.html JS label() 同格式：姓名（電話） / 姓名）。"""
    if not cust:
        return ""
    return f"{cust.name}（{cust.phone}）" if cust.phone else cust.name


def _prefill_customer_form(db, customer_id_raw):
    """CR-7：?customer_id= 預填客戶欄；無 / 不存在 ⇒ 空 form。"""
    cid = _to_int(customer_id_raw)
    cust = db.get(Customer, cid) if cid else None
    if not cust:
        return {}
    return {"customer_id": str(cust.id), "customer_display": customer_display(cust)}


def _parse_shipping_fields():
    """CR-5：運費 / 運送方式 / 運送備註（皆選填，預設 0 / 空）。

    回 (fields_dict, error)。運送方式非閉集 ⇒ error；運費負數視為 0。
    """
    fee = _to_decimal(request.form.get("shipping_fee"), default=0)
    if fee < 0:
        fee = Decimal("0")
    method = (request.form.get("shipping_method") or "").strip() or None
    if method is not None and method not in SHIPPING_METHODS:
        return None, f"無效的運送方式：{method}"
    ship_note = (request.form.get("shipping_note") or "").strip() or None
    if ship_note and len(ship_note) > 256:
        return None, "運送備註過長（上限 256 字）"
    return {"shipping_fee": fee, "shipping_method": method,
            "shipping_note": ship_note}, None


def _create_order(db):
    operator = _operator_name()

    # ---- 1. 解析表單 ----
    customer_id = _to_int(request.form.get("customer_id"))
    # 建單當場新增客戶（森哥 2026-08-19）：客戶欄可打字過濾，打的名字不在名單時
    # 前端把它放進 new_customer_name。此處在「同一張訂單的 transaction」內建客戶，
    # 訂單若整張 rollback，客戶也不會留下半筆。已選既有客戶時以 customer_id 為準。
    new_customer_name = (request.form.get("new_customer_name") or "").strip()
    if customer_id:
        new_customer_name = ""
    if len(new_customer_name) > 64:
        flash("新客戶姓名過長（上限 64 字），訂單未建立")
        return _redraw_form(db)
    recipient_name = (request.form.get("recipient_name") or "").strip()
    recipient_phone = (request.form.get("recipient_phone") or "").strip()
    shipping_address = (request.form.get("shipping_address") or "").strip()
    note = (request.form.get("note") or "").strip() or None
    discount = _to_decimal(request.form.get("discount"), default=0)
    if discount < 0:
        discount = Decimal("0")
    ship_fields, ship_err = _parse_shipping_fields()
    if ship_err:
        flash(ship_err)
        return _redraw_form(db)

    rows, item_err = _parse_items(db)
    if item_err:
        flash(item_err)
        return _redraw_form(db)

    total_amount = _calc_total(rows, discount)

    # ---- 2. 整張原子 transaction（§4.2）----
    created_customer = None
    if new_customer_name:
        existing = (db.query(Customer)
                      .filter(Customer.name == new_customer_name).first())
        if existing is not None:
            customer_id = existing.id      # 同名已存在就沿用，不重複建一筆
        else:
            cust = Customer(
                name=new_customer_name,
                phone=recipient_phone or None,
                created_by=_uid(), updated_by=_uid(),
            )
            db.add(cust)
            db.flush()
            customer_id = cust.id
            created_customer = new_customer_name

    order_no = _gen_order_no(db)
    order = Order(
        order_no=order_no,
        customer_id=customer_id,
        recipient_name=recipient_name or None,
        recipient_phone=recipient_phone or None,
        shipping_address=shipping_address or None,
        total_amount=total_amount,      # 商品實收 − 折扣（不含運費）
        discount=discount,
        shipping_fee=ship_fields["shipping_fee"],
        shipping_method=ship_fields["shipping_method"],
        shipping_note=ship_fields["shipping_note"],
        payment_status="unpaid",
        shipping_status="pending",
        note=note,
        created_by=_uid(),
        updated_by=_uid(),
    )
    db.add(order)
    db.flush()  # 取得 order.id，供 order_ref 與明細外鍵使用

    order_ref = str(order.id)
    try:
        for r in rows:
            db.add(OrderItem(
                order_id=order.id,
                product_id=r["product_id"],
                combo_code=r["combo_code"],
                qty=r["qty"],
                unit_price=r["unit_price"],
                subtotal=r["subtotal"],
            ))
        db.flush()

        # 逐項扣庫存：呼叫 §4.1 契約函式（同一 session / 同一 tx）
        for r in rows:
            result = inventory_service.deduct_for_sale(
                db,
                product_id=r["product_id"],
                combo_code=r["combo_code"],
                order_qty=r["qty"],
                operator=operator,
                order_ref=order_ref,
                note=f"order {order_no}",
            )
            if not _result_ok(result):
                # 任一品項缺貨 ⇒ 整張 rollback（§4.2）
                db.rollback()
                pname = _product_label(db, r["product_id"])
                flash(f"庫存不足，整張訂單未建立：{pname} / {combo_label(r['combo_code'])} × {r['qty']}")
                return _redraw_form(db)

        # CR-8：建單留痕（同一 tx；整張 rollback 時一併消失）
        _audit(db, "order_create", order.id,
               _order_create_detail(db, order, rows, created_customer))
        db.commit()
    except Exception as exc:  # noqa: BLE001 — 任何異動失敗整張 rollback（R1）
        db.rollback()
        flash(f"建立訂單失敗，已全數回復：{exc}")
        return _redraw_form(db)

    if created_customer:
        flash(f"訂單 {order_no} 已建立，庫存已扣減；同時新增客戶「{created_customer}」")
    else:
        flash(f"訂單 {order_no} 已建立，庫存已扣減")
    return redirect(url_for("orders.detail", order_id=order.id))


# -------------------------------------------------------------------------
# 改付款狀態
# -------------------------------------------------------------------------
@orders_bp.route("/<int:order_id>/payment", methods=["POST"])
@role_required(*WRITE_ROLES)
def update_payment(order_id):
    db = get_session()
    order = db.get(Order, order_id)
    if not order:
        abort(404)
    if order.voided_at is not None:
        flash("此訂單已作廢，不可再改狀態")
        return redirect(url_for("orders.detail", order_id=order_id))
    new_status = (request.form.get("payment_status") or "").strip()
    if new_status not in PAYMENT_STATUSES:
        flash("無效的付款狀態")
        return redirect(url_for("orders.detail", order_id=order_id))
    old_status = order.payment_status
    order.payment_status = new_status
    order.updated_by = _uid()
    if old_status != new_status:   # CR-8：付款狀態變更留痕（before → after）
        _audit(db, "order_payment_status", order.id, {
            "order_no": order.order_no,
            "before": {"payment_status": old_status},
            "after": {"payment_status": new_status},
        })
    db.commit()
    flash("付款狀態已更新")
    return redirect(url_for("orders.detail", order_id=order_id))


# -------------------------------------------------------------------------
# 改出貨狀態（出貨成立同一 tx 扣紙袋，§4.4）
# -------------------------------------------------------------------------
@orders_bp.route("/<int:order_id>/shipping", methods=["POST"])
@role_required(*WRITE_ROLES)
def update_shipping(order_id):
    db = get_session()
    order = db.get(Order, order_id)
    if not order:
        abort(404)
    if order.voided_at is not None:
        flash("此訂單已作廢，不可再改狀態")
        return redirect(url_for("orders.detail", order_id=order_id))
    new_status = (request.form.get("shipping_status") or "").strip()
    if new_status not in SHIPPING_STATUSES:
        flash("無效的出貨狀態（取消訂單請改用「作廢」）")
        return redirect(url_for("orders.detail", order_id=order_id))

    old_status = order.shipping_status
    operator = _operator_name()

    # 由「未出貨」→「shipped」才觸發紙袋扣減（§4.4），且避免重複扣
    triggers_shipment = (new_status == "shipped" and old_status != "shipped")

    try:
        order.shipping_status = new_status
        order.updated_by = _uid()
        db.flush()

        if triggers_shipment:
            result = inventory_service.deduct_for_shipment(
                db, order_id=order.id, operator=operator
            )
            if not _result_ok(result):
                # 紙袋不足 ⇒ rollback、出貨不成立（§4.4）
                db.rollback()
                flash("紙袋庫存不足，出貨未成立（已回復）")
                return redirect(url_for("orders.detail", order_id=order_id))
            # CR-4：桌面出貨也補建 shipments 列（原只有手機 ship_order 建），
            # 使「已出貨」判定＝shipments 有列；tracking/carrier 選填
            db.add(Shipment(
                order_id=order.id,
                tracking_no=(request.form.get("tracking_no") or "").strip() or None,
                carrier=(request.form.get("carrier") or "").strip() or None,
                shipped_at=datetime.utcnow(),
                status="shipped",
                created_by=_uid(),
            ))

        if old_status != new_status:   # CR-8：出貨狀態變更留痕（before → after）
            _audit(db, "order_shipping_status", order.id, {
                "order_no": order.order_no,
                "before": {"shipping_status": old_status},
                "after": {"shipping_status": new_status},
                "paper_bag_deducted": triggers_shipment,
                "tracking_no": (request.form.get("tracking_no") or "").strip() or None,
                "carrier": (request.form.get("carrier") or "").strip() or None,
            })
        db.commit()
    except Exception as exc:  # noqa: BLE001
        db.rollback()
        flash(f"更新出貨狀態失敗，已回復：{exc}")
        return redirect(url_for("orders.detail", order_id=order_id))

    flash("出貨狀態已更新" + ("（已扣紙袋）" if triggers_shipment else ""))
    return redirect(url_for("orders.detail", order_id=order_id))


# -------------------------------------------------------------------------
# 歷史訂單匯入（D14 完整實作）
#
# 匯入是「歷史資料載入」（歷史已發生）：
#   - 不扣庫存、不走 inventory_service、不寫 inventory_movements / inventory_balances。
#   - 只落地到 historical_order_imports（批次）+ historical_order_import_rows（逐列）。
#   - 每列保留原始欄位（raw_json）+ error_message；統計 imported/success/failed。
# 權限限 owner / accounting（§六 finance/歷史資料；owner 永遠通過）。
# -------------------------------------------------------------------------

# D14 欄位對應（藍本）。表頭比對採「中文名稱 或 欄位代碼」皆可命中（去空白）。
IMPORT_FIELDS = [
    ("order_no", "訂單編號"),
    ("name", "姓名"),
    ("phone", "電話"),
    ("order_date", "訂購日期"),
    ("product_name", "商品名稱"),
    ("qty", "購買數量"),
    ("unit_price", "單價"),
    ("amount", "訂單金額"),
    ("payment_status", "付款狀態"),
    ("shipping_status", "出貨狀態"),
    ("note", "備註"),
]
IMPORT_WRITE_ROLES = ("accounting",)  # owner 由 role_required 永遠通過


@orders_bp.route("/import", methods=["GET", "POST"])
@role_required(*IMPORT_WRITE_ROLES)
def import_page():
    db = get_session()
    if request.method == "POST":
        return _do_import(db)
    # GET：上傳頁 + 欄位對照 + 歷史批次列表
    batches = (
        db.query(HistoricalOrderImport)
        .order_by(HistoricalOrderImport.id.desc())
        .limit(50)
        .all()
    )
    return render_template(
        "orders/import.html", section="orders",
        import_fields=IMPORT_FIELDS, batches=batches,
    )


@orders_bp.route("/import/<int:import_id>")
@role_required(*IMPORT_WRITE_ROLES)
def import_result(import_id):
    db = get_session()
    batch = db.get(HistoricalOrderImport, import_id)
    if not batch:
        abort(404)
    rows = (
        db.query(HistoricalOrderImportRow)
        .filter_by(import_id=import_id)
        .order_by(HistoricalOrderImportRow.id.asc())
        .all()
    )
    return render_template(
        "orders/import_result.html", section="orders",
        batch=batch, rows=rows, import_fields=IMPORT_FIELDS,
    )


# ---- 匯入核心 ------------------------------------------------------------

# 表頭文字 -> 內部欄位代碼（中文名 + 英文碼皆可命中）
_HEADER_ALIAS = {}
for _code, _label in IMPORT_FIELDS:
    _HEADER_ALIAS[_code] = _code
    _HEADER_ALIAS[_label] = _code


def _norm_header(h):
    return ("" if h is None else str(h)).strip()


def _parse_rows_from_file(filename, raw_bytes):
    """解析 Excel/CSV → (headers, list[dict])。dict 以原始表頭文字為 key。

    回傳的 list 每筆是 {原始表頭: 值}；後續再依 _HEADER_ALIAS 對應內部欄位。
    解析失敗（檔案層級）拋 ValueError。
    """
    name = (filename or "").lower()
    records = []
    if name.endswith(".csv") or name.endswith(".txt"):
        text_data = None
        for enc in ("utf-8-sig", "utf-8", "big5", "cp950"):
            try:
                text_data = raw_bytes.decode(enc)
                break
            except UnicodeDecodeError:
                continue
        if text_data is None:
            raise ValueError("CSV 編碼無法辨識（試過 utf-8/big5）")
        reader = csv.reader(io.StringIO(text_data))
        all_rows = [r for r in reader]
        if not all_rows:
            return [], []
        headers = [_norm_header(h) for h in all_rows[0]]
        for r in all_rows[1:]:
            rec = {}
            for i, h in enumerate(headers):
                rec[h] = r[i] if i < len(r) else None
            records.append(rec)
        return headers, records

    if name.endswith(".xlsx") or name.endswith(".xlsm"):
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            return [], []
        headers = [_norm_header(h) for h in header_row]
        for r in rows_iter:
            if r is None:
                continue
            rec = {}
            for i, h in enumerate(headers):
                rec[h] = r[i] if i < len(r) else None
            records.append(rec)
        return headers, records

    raise ValueError("僅支援 .xlsx / .csv 檔")


def _map_record(rec):
    """將原始列（以表頭為 key）映射到內部欄位代碼字典。未命中的表頭忽略。"""
    mapped = {}
    for raw_header, val in rec.items():
        code = _HEADER_ALIAS.get(_norm_header(raw_header))
        if code:
            mapped[code] = val
    return mapped


def _to_dec_or_none(v):
    if v is None or str(v).strip() == "":
        return None
    try:
        return Decimal(str(v).strip())
    except (InvalidOperation, ValueError):
        raise ValueError(f"無法解析數值：{v!r}")


def _to_int_or_none(v):
    if v is None or str(v).strip() == "":
        return None
    try:
        return int(float(str(v).strip()))
    except (TypeError, ValueError):
        raise ValueError(f"無法解析整數：{v!r}")


def _validate_and_build_row(mapped, raw_rec):
    """從 mapped 內部欄位建一筆 HistoricalOrderImportRow（不 commit）。

    回傳 (row_obj, is_success)。解析錯誤時 error_message 非空、is_success=False，
    但仍落地該列（保留 raw + 錯誤），符合 D14「逐列寫、保留 raw + error_message」。
    """
    errors = []
    qty = unit_price = amount = None
    try:
        qty = _to_int_or_none(mapped.get("qty"))
    except ValueError as e:
        errors.append(str(e))
    try:
        unit_price = _to_dec_or_none(mapped.get("unit_price"))
    except ValueError as e:
        errors.append(str(e))
    try:
        amount = _to_dec_or_none(mapped.get("amount"))
    except ValueError as e:
        errors.append(str(e))

    # 必要欄位輕量檢查：訂單編號 + 商品名稱其一缺 → 記為失敗（仍落地）
    if not (mapped.get("order_no") and str(mapped.get("order_no")).strip()):
        errors.append("缺訂單編號")
    if not (mapped.get("product_name") and str(mapped.get("product_name")).strip()):
        errors.append("缺商品名稱")

    err_msg = "；".join(errors) if errors else None

    def _s(key):
        v = mapped.get(key)
        return None if v is None else str(v).strip() or None

    row = HistoricalOrderImportRow(
        order_no=_s("order_no"),
        name=_s("name"),
        phone=_s("phone"),
        order_date=_s("order_date"),
        product_name=_s("product_name"),
        qty=qty,
        unit_price=unit_price,
        amount=amount,
        payment_status=_s("payment_status"),
        shipping_status=_s("shipping_status"),
        note=_s("note"),
        raw_json=json.dumps(
            {str(k): (None if v is None else str(v)) for k, v in raw_rec.items()},
            ensure_ascii=False,
        ),
        error_message=err_msg,
    )
    return row, (err_msg is None)


def _do_import(db):
    f = request.files.get("file")
    if not f or not f.filename:
        flash("請選擇要匯入的 Excel/CSV 檔")
        return redirect(url_for("orders.import_page"))

    raw_bytes = f.read()
    try:
        headers, records = _parse_rows_from_file(f.filename, raw_bytes)
    except ValueError as e:
        flash(f"匯入失敗：{e}")
        return redirect(url_for("orders.import_page"))
    except Exception as e:  # noqa: BLE001
        flash(f"匯入失敗（解析錯誤）：{e}")
        return redirect(url_for("orders.import_page"))

    batch_name = (request.form.get("batch_name") or "").strip() \
        or f"匯入 {datetime.utcnow().strftime('%Y%m%d-%H%M%S')}"

    # ---- 落地：批次 + 逐列（單一 transaction；不碰庫存）----
    try:
        batch = HistoricalOrderImport(
            batch_name=batch_name,
            source_file=f.filename,
            row_count=len(records),
            imported_rows=0,
            success_rows=0,
            failed_rows=0,
            status="pending",
            imported_by=_uid(),
        )
        db.add(batch)
        db.flush()  # 取得 batch.id

        success = 0
        failed = 0
        for rec in records:
            mapped = _map_record(rec)
            row, ok = _validate_and_build_row(mapped, rec)
            row.import_id = batch.id
            db.add(row)
            if ok:
                success += 1
            else:
                failed += 1

        batch.imported_rows = len(records)
        batch.success_rows = success
        batch.failed_rows = failed
        batch.status = "completed" if failed == 0 else "completed_with_errors"
        db.commit()
    except Exception as e:  # noqa: BLE001
        db.rollback()
        flash(f"匯入寫入失敗，已全數回復：{e}")
        return redirect(url_for("orders.import_page"))

    flash(f"匯入完成：{batch_name}（共 {len(records)} 列，成功 {success}，失敗 {failed}）")
    return redirect(url_for("orders.import_result", import_id=batch.id))


# -------------------------------------------------------------------------
# helpers
# -------------------------------------------------------------------------
def _uid():
    u = current_user()
    return u.id if u else None


def _operator_name():
    u = current_user()
    if not u:
        return "system"
    return u.display_name or u.username


def _redraw_form(db):
    return render_template(
        "orders/form.html", section="orders",
        customers=db.query(Customer).order_by(Customer.name.asc()).all(),
        products=sellable_products(db),
        unit_codes=UNIT_CODES, shipping_methods=SHIPPING_METHODS,
        form=request.form,
    )


def _gen_order_no(db):
    """產生唯一訂單編號 FC + yyyymmdd + 當日序號（碰撞則遞增）。"""
    today = datetime.utcnow().strftime("%Y%m%d")
    prefix = f"FC{today}"
    last = (
        db.query(Order)
        .filter(Order.order_no.like(f"{prefix}%"))
        .order_by(Order.order_no.desc())
        .first()
    )
    seq = 1
    if last and last.order_no.startswith(prefix):
        try:
            seq = int(last.order_no[len(prefix):]) + 1
        except ValueError:
            seq = 1
    candidate = f"{prefix}{seq:03d}"
    # 防併發碰撞：若已存在則往後找
    while db.query(Order).filter_by(order_no=candidate).first():
        seq += 1
        candidate = f"{prefix}{seq:03d}"
    return candidate


def _product_label(db, product_id):
    p = db.get(Product, product_id)
    return p.name if p else f"#{product_id}"


def _to_int(v):
    try:
        return int(v)
    except (TypeError, ValueError):
        return None


def _to_decimal(v, default=0):
    from decimal import Decimal, InvalidOperation
    if v is None or str(v).strip() == "":
        return Decimal(str(default))
    try:
        return Decimal(str(v).strip())
    except (InvalidOperation, ValueError):
        return Decimal(str(default))


def _result_ok(result):
    """容忍 inventory_service 回傳型別：物件(.ok) / dict(['ok']) / bool / None。

    None 視為未表態 → 為安全起見當作失敗；契約規定缺貨須回 ok=false（§4.1）。
    """
    if result is None:
        return False
    if isinstance(result, bool):
        return result
    if isinstance(result, dict):
        return bool(result.get("ok"))
    ok = getattr(result, "ok", None)
    if ok is None:
        # 物件存在但無 ok 欄位 → 視為成功（service 以例外表達失敗的情況）
        return True
    return bool(ok)


# =========================================================================
# CR-4（2026-08-31）訂單編輯 / 作廢
# =========================================================================
def has_shipment_record(db, order_id):
    """shipments 表是否有該單列（桌面 update_shipping 轉 shipped／手機 ship_order 皆建）。"""
    return db.query(Shipment.id).filter_by(order_id=order_id).first() is not None


def is_order_shipped(db, order_id):
    """「已出貨」判定（CR-10 改規則）：**只看 shipments 表有沒有該單列**。

    有列 = 真正走過出貨流程（可能已扣紙袋）→ 編輯時品項鎖定、作廢不回補且限 owner。
    無列 = 視為未出貨，即使 shipping_status 已是 shipped/delivered（客戶直接以 delivered 建單的歷史單、
    手機「僅改狀態旗標」的單）→ 可正常反沖→重建→重扣、可回補作廢。
    舊規則（狀態 shipped/delivered 亦鎖）已移除：線上全部訂單都被鎖死，加行／刪行做不到。
    """
    return has_shipment_record(db, order_id)


def sellable_products(db):
    """建單／編輯頁商品下拉：上架面膜商品 + 上架包材商品（紙袋當品項，2026-09-01）。面膜排前、包材排後。"""
    prods = db.query(Product).filter_by(active=True).order_by(Product.id.asc()).all()
    return [p for p in prods if not p.is_packaging] + [p for p in prods if p.is_packaging]


def validate_row_units(db, rows):
    """商品 × 單位 合法性（桌面／手機建單與編輯共用）：包材商品只能 BAG(袋)，面膜商品只能 LOOSE/BOX。

    rows：dict list，至少含 product_id / combo_code。回錯誤字串或 None。
    """
    ids = {int(r["product_id"]) for r in rows}
    prods = {p.id: p for p in db.query(Product).filter(Product.id.in_(ids)).all()} if ids else {}
    for r in rows:
        p = prods.get(int(r["product_id"]))
        if p is None or not p.active:
            return f"商品不存在或已下架：#{r['product_id']}"
        combo = r["combo_code"]
        if p.is_packaging and combo not in BAG_UNIT_CODES:
            return f"{p.name} 為包材，單位只能選「袋」"
        if not p.is_packaging and combo in BAG_UNIT_CODES:
            return f"{p.name} 不是包材，不能以「袋」為單位"
    return None


def _parse_items(db=None):
    """解析多品項平行陣列 → (rows, error)。建單與編輯共用。

    combo_codes 存單位值 LOOSE/BOX/BAG；amounts = 該列實收金額（直接金額，非單價×數量）；
    unit_price 與 subtotal 同存此金額（無 schema 變更）。
    有給 db 時另做 validate_row_units（包材只能袋、面膜只能片/盒）。
    """
    product_ids = request.form.getlist("item_product_id")
    combo_codes = request.form.getlist("item_combo_code")
    qtys = request.form.getlist("item_qty")
    amounts = request.form.getlist("item_amount")

    rows = []
    for i in range(len(product_ids)):
        pid = _to_int(product_ids[i])
        combo = (combo_codes[i] if i < len(combo_codes) else "").strip()
        qty = _to_int(qtys[i] if i < len(qtys) else None)
        amount = _to_decimal(amounts[i] if i < len(amounts) else None, default=0)
        if not pid or not combo:
            continue
        if combo not in UNIT_CODES:
            return None, f"無效的品項單位：{combo}（限 片 / 盒 / 袋）"
        if not qty or qty <= 0:
            return None, "品項數量必須為正整數"
        rows.append({"product_id": pid, "combo_code": combo, "qty": qty,
                     "unit_price": amount, "subtotal": amount})
    if not rows:
        return None, "訂單至少需一個品項"
    if db is not None:
        unit_err = validate_row_units(db, rows)
        if unit_err:
            return None, unit_err
    return rows, None


def _calc_total(rows, discount):
    """total_amount = Σsubtotal − 折扣（不含運費，CR-5 口徑 B）；下限 0。"""
    total = sum(r["subtotal"] for r in rows) - discount
    return total if total > 0 else Decimal("0")


def _json_default(v):
    if isinstance(v, Decimal):
        return str(v)
    if isinstance(v, datetime):
        return v.isoformat()
    return str(v)


def _order_snapshot(db, order):
    """audit before/after 快照（items + 金額 + 收件 + 運送）。"""
    items = db.query(OrderItem).filter_by(order_id=order.id).order_by(OrderItem.id.asc()).all()
    return {
        "order_no": order.order_no,
        "customer_id": order.customer_id,
        "recipient_name": order.recipient_name,
        "recipient_phone": order.recipient_phone,
        "shipping_address": order.shipping_address,
        "note": order.note,
        "discount": order.discount,
        "shipping_fee": order.shipping_fee,
        "shipping_method": order.shipping_method,
        "shipping_note": order.shipping_note,
        "total_amount": order.total_amount,
        "payment_status": order.payment_status,
        "shipping_status": order.shipping_status,
        "items": [{"product_id": it.product_id, "combo_code": it.combo_code,
                   "qty": it.qty, "subtotal": it.subtotal} for it in items],
    }


def _audit(db, action, order_id, detail):
    """訂單類留痕（CR-8 起走共用 audit_util.write_audit；actor 取 session 使用者）。"""
    write_audit(db, action, "orders", order_id, detail)


def _order_create_detail(db, order, rows, created_customer=None):
    """建單留痕 detail：單號 / 客戶 / 品項摘要 / 金額（CR-8）。rows 為 (product_id, combo_code, qty, unit_price, subtotal) dict。"""
    cust_name = None
    if order.customer_id:
        c = db.get(Customer, order.customer_id)
        cust_name = c.name if c else None
    return {
        "order_no": order.order_no,
        "customer_id": order.customer_id, "customer_name": cust_name,
        "recipient_name": order.recipient_name, "recipient_phone": order.recipient_phone,
        "items": [{
            "product_id": r["product_id"], "product_name": _product_label(db, r["product_id"]),
            "combo_code": r["combo_code"], "qty": r["qty"],
            "unit_price": r.get("unit_price"), "subtotal": r.get("subtotal"),
        } for r in rows],
        "total_amount": order.total_amount, "discount": order.discount,
        "shipping_fee": order.shipping_fee, "shipping_method": order.shipping_method,
        "created_customer": created_customer,
    }


def _resolve_customer(db, customer_id, new_customer_name, recipient_phone):
    """客戶欄：已選既有客戶以其為準；否則打的名字不在名單就當場建（同 tx）。回 (customer_id, created_name)。"""
    if customer_id or not new_customer_name:
        return customer_id, None
    existing = db.query(Customer).filter(Customer.name == new_customer_name).first()
    if existing is not None:
        return existing.id, None
    cust = Customer(name=new_customer_name, phone=recipient_phone or None,
                    created_by=_uid(), updated_by=_uid())
    db.add(cust)
    db.flush()
    return cust.id, new_customer_name


def _edit_form_dict(db, order):
    """GET 編輯頁：以訂單現值預填 form.html。"""
    cust = db.get(Customer, order.customer_id) if order.customer_id else None
    return {
        "customer_id": str(order.customer_id or ""),
        "customer_display": customer_display(cust),
        "recipient_name": order.recipient_name or "",
        "recipient_phone": order.recipient_phone or "",
        "shipping_address": order.shipping_address or "",
        "discount": str(order.discount or 0),
        "shipping_fee": str(order.shipping_fee or 0),
        "shipping_method": order.shipping_method or "",
        "shipping_note": order.shipping_note or "",
        "note": order.note or "",
    }


def _items_from_form():
    """POST 驗證失敗重繪編輯頁時，品項列以剛送出的表單為準（不是 DB 舊值），使用者不必重打。"""
    pids = request.form.getlist("item_product_id")
    combos = request.form.getlist("item_combo_code")
    qtys = request.form.getlist("item_qty")
    amts = request.form.getlist("item_amount")
    out = []
    for i, pid in enumerate(pids):
        out.append({
            "product_id": _to_int(pid),
            "combo_code": (combos[i] if i < len(combos) else ""),
            "qty": _to_int(qtys[i] if i < len(qtys) else None) or 1,
            "amount": (amts[i] if i < len(amts) else "0") or "0",
        })
    return out


def _render_edit(db, order, form=None):
    from_form = form is not None
    items = db.query(OrderItem).filter_by(order_id=order.id).order_by(OrderItem.id.asc()).all()
    edit_items = (_items_from_form() if from_form else
                  [{"product_id": it.product_id, "combo_code": it.combo_code,
                    "qty": it.qty, "amount": str(it.subtotal or 0)} for it in items])
    return render_template(
        "orders/form.html", section="orders",
        customers=db.query(Customer).order_by(Customer.name.asc()).all(),
        products=sellable_products(db),
        unit_codes=UNIT_CODES, shipping_methods=SHIPPING_METHODS,
        form=(form if from_form else _edit_form_dict(db, order)),
        edit_order=order,
        edit_items=edit_items,
        items_locked=is_order_shipped(db, order.id),
        form_action=url_for("orders.edit", order_id=order.id),
    )


def _lines_diff(db, before_items, after_rows):
    """CR-10：比對編輯前後品項列，回 (lines_added, lines_removed, lines_changed)。

    以 (product_id, combo_code) 為列身分、同身分依出現順序配對：
      只在後 → 新增；只在前 → 刪除；兩邊都有但 qty 或 subtotal 不同 → 修改（含 before/after）。
    每列附 product_name，供操作紀錄頁人話摘要。
    """
    def key(r):
        return (int(r["product_id"]), str(r["combo_code"]))

    def norm_amt(v):
        try:
            return Decimal(str(v if v is not None else 0))
        except (InvalidOperation, ValueError):
            return Decimal("0")

    def line(r):
        return {"product_id": r["product_id"],
                "product_name": _product_label(db, r["product_id"]),
                "combo_code": r["combo_code"], "qty": r["qty"],
                "subtotal": norm_amt(r.get("subtotal"))}

    before_by = {}
    for r in before_items:
        before_by.setdefault(key(r), []).append(r)
    added, removed, changed = [], [], []
    for r in after_rows:
        pool = before_by.get(key(r))
        if pool:
            b = pool.pop(0)
            if int(b["qty"]) != int(r["qty"]) or norm_amt(b.get("subtotal")) != norm_amt(r.get("subtotal")):
                changed.append({**line(r), "qty_before": b["qty"], "qty_after": r["qty"],
                                "subtotal_before": norm_amt(b.get("subtotal")),
                                "subtotal_after": norm_amt(r.get("subtotal"))})
        else:
            added.append(line(r))
    for pool in before_by.values():
        for b in pool:
            removed.append(line(b))
    return added, removed, changed


def perform_edit(db, order, rows, discount, fields, operator, actor_id, via=None):
    """訂單編輯核心（CR-10 抽出；桌面 edit / 手機 order_edit 共用）。同一 tx，呼叫端 commit/rollback。

    rows     ：新品項列（_parse_items 格式 dict list）；品項鎖定（有出貨紀錄）時傳 None → 不動 items / 庫存 / 折扣。
    discount ：新折扣（Decimal）；rows 為 None 時忽略。
    fields   ：要更新的欄位 dict（customer_id / recipient_name / recipient_phone / shipping_address /
               note / shipping_fee / shipping_method / shipping_note），只更新有給的鍵；None 或 {} = 不動。
    流程（R1）：reverse_sale 整單反沖 → 刪舊 order_items → 重建 → 逐列 deduct_for_sale → 重算 total
              → 更新欄位 → AuditLog(order_edit，含 lines_added/removed/changed)。
    回 (ok, message)；失敗時呼叫端 rollback（本函式不 rollback、不 commit）。
    """
    locked = is_order_shipped(db, order.id)
    if locked:
        rows = None
    before = _order_snapshot(db, order)
    reversal_ids, sale_ids = [], []
    added, removed, changed = [], [], []

    if rows is not None:
        rv = inventory_service.reverse_sale(
            db, order.id, operator=operator, actor_id=actor_id, reason="order_edit")
        if not _result_ok(rv):
            return False, f"編輯失敗（庫存回補被拒）：{getattr(rv, 'error', '')}"
        reversal_ids = list(getattr(rv, "movement_ids", []) or [])

        added, removed, changed = _lines_diff(db, before["items"], rows)
        db.query(OrderItem).filter_by(order_id=order.id).delete(synchronize_session=False)
        db.flush()
        for r in rows:
            db.add(OrderItem(
                order_id=order.id, product_id=r["product_id"],
                combo_code=r["combo_code"], qty=r["qty"],
                unit_price=r["unit_price"], subtotal=r["subtotal"],
            ))
        db.flush()
        for r in rows:
            result = inventory_service.deduct_for_sale(
                db, product_id=r["product_id"], combo_code=r["combo_code"],
                order_qty=r["qty"], operator=operator, order_ref=str(order.id),
                note=f"order {order.order_no}（編輯重扣）",
            )
            if not _result_ok(result):
                # 任一缺貨 ⇒ 呼叫端整張 rollback（含反沖），餘量不變
                pname = _product_label(db, r["product_id"])
                return False, (f"庫存不足，編輯未儲存（已全數回復）：{pname} / "
                               f"{combo_label(r['combo_code'])} × {r['qty']}")
            sale_ids.extend(getattr(result, "movement_ids", []) or [])
        if discount is None or discount < 0:
            discount = Decimal("0")
        order.discount = discount
        order.total_amount = _calc_total(rows, discount)

    for k, v in (fields or {}).items():
        setattr(order, k, v)
    order.updated_by = actor_id
    db.flush()

    after = _order_snapshot(db, order)
    detail = {
        "before": before, "after": after, "items_locked": locked,
        "lines_added": added, "lines_removed": removed, "lines_changed": changed,
        "reversal_movement_ids": reversal_ids, "sale_movement_ids": sale_ids,
    }
    if via:
        detail["via"] = via
    _audit(db, "order_edit", order.id, detail)
    msg = "訂單已更新" + ("（有出貨紀錄：品項鎖定，僅更新收件/運費/備註）" if locked
                          else "（庫存已依新品項重算）")
    return True, msg


@orders_bp.route("/<int:order_id>/edit", methods=["GET", "POST"])
@role_required(*WRITE_ROLES)
def edit(order_id):
    """編輯訂單（staff + owner）。核心走 perform_edit（CR-10 抽出，手機共用）。

    未出貨（shipments 無列）：同一 tx「reverse_sale 反沖 → 刪舊 order_items → 依表單重建（可新增／刪除列）
            → 逐列 deduct_for_sale → 重算 total_amount → 更新收件/備註/運費/方式 → AuditLog(order_edit)」；
            任一缺貨整張 rollback。
    有出貨紀錄（shipments 有列）：品項與折扣鎖定，只改備註／運費／運送方式／收件資訊，不動庫存。
    """
    db = get_session()
    order = db.get(Order, order_id)
    if not order:
        abort(404)
    if order.voided_at is not None:
        flash("此訂單已作廢，不可編輯")
        return redirect(url_for("orders.detail", order_id=order_id))

    if request.method == "GET":
        return _render_edit(db, order)

    locked = is_order_shipped(db, order_id)

    # ---- 1. 解析表單 ----
    # 客戶欄：表單有帶 customer_id（含空字串＝明確清除）才改；完全沒帶（簡版／程式化 POST）則保留原客戶，
    # 避免一次品項編輯把訂單從客戶名單裡弄丟（CR-10）。
    customer_id = (_to_int(request.form.get("customer_id"))
                   if "customer_id" in request.form else order.customer_id)
    new_customer_name = (request.form.get("new_customer_name") or "").strip()
    if customer_id:
        new_customer_name = ""
    if len(new_customer_name) > 64:
        flash("新客戶姓名過長（上限 64 字），未儲存")
        return _render_edit(db, order, form=request.form)
    recipient_name = (request.form.get("recipient_name") or "").strip()
    recipient_phone = (request.form.get("recipient_phone") or "").strip()
    shipping_address = (request.form.get("shipping_address") or "").strip()
    note = (request.form.get("note") or "").strip() or None
    ship_fields, ship_err = _parse_shipping_fields()
    if ship_err:
        flash(ship_err)
        return _render_edit(db, order, form=request.form)

    rows = None
    discount = None
    if not locked:
        discount = _to_decimal(request.form.get("discount"), default=0)
        rows, item_err = _parse_items(db)
        if item_err:
            flash(item_err)
            return _render_edit(db, order, form=request.form)

    # ---- 2. 同一原子 tx（R1）----
    try:
        customer_id, _created = _resolve_customer(
            db, customer_id, new_customer_name, recipient_phone)
        ok, msg = perform_edit(
            db, order, rows, discount,
            fields={
                "customer_id": customer_id,
                "recipient_name": recipient_name or None,
                "recipient_phone": recipient_phone or None,
                "shipping_address": shipping_address or None,
                "note": note,
                "shipping_fee": ship_fields["shipping_fee"],
                "shipping_method": ship_fields["shipping_method"],
                "shipping_note": ship_fields["shipping_note"],
            },
            operator=_operator_name(), actor_id=_uid(),
        )
        if not ok:
            db.rollback()
            flash(msg)
            return redirect(url_for("orders.edit", order_id=order_id))
        db.commit()
    except Exception as exc:  # noqa: BLE001 — 任何異動失敗整張 rollback（R1）
        db.rollback()
        flash(f"編輯失敗，已全數回復：{exc}")
        return redirect(url_for("orders.edit", order_id=order_id))

    flash(msg)
    return redirect(url_for("orders.detail", order_id=order_id))


def perform_void(db, order, reason, operator, actor_id, actor_name, extra_detail=None):
    """作廢核心（桌面 / 手機 / 批次共用；權限檢查在路由層）。同一 tx，呼叫端 commit/rollback。

    - 軟刪除：設 voided_at / voided_by / void_reason；不刪 payments / shipments / order_items。
    - 未出貨 → reverse_sale 全額回補（normal 池，不碰 reserved）。
    - 已出貨（shipments 有列）→ 不回補（貨與袋已出門，退貨另走 RESTOCK 入庫留痕）。
    - payment_status='paid' → 改 'refunded'。
    - AuditLog(order_void, detail 含 items 快照 / total / movement ids；extra_detail 併入，
      CR-6 批次用 bulk=True / bulk_group_id）。
    回 (ok, message)。
    """
    if order.voided_at is not None:
        return False, "此訂單已作廢"
    reason = (reason or "").strip()
    if not reason:
        return False, "作廢必須填寫原因"
    shipped = is_order_shipped(db, order.id)
    snapshot = _order_snapshot(db, order)
    mv_ids = []
    if not shipped:
        rv = inventory_service.reverse_sale(
            db, order.id, operator=operator, actor_id=actor_id, reason="order_void")
        if not _result_ok(rv):
            return False, f"庫存回補被拒，未作廢：{getattr(rv, 'error', '')}"
        mv_ids = list(getattr(rv, "movement_ids", []) or [])

    old_pay = order.payment_status
    if old_pay == "paid":
        order.payment_status = "refunded"
    order.voided_at = datetime.utcnow()
    order.voided_by = actor_id
    order.void_reason = reason[:256]
    order.updated_by = actor_id
    db.flush()
    detail = {
        "reason": reason, "shipped": shipped, "stock_reversed": (not shipped),
        "payment_status_before": old_pay, "payment_status_after": order.payment_status,
        "snapshot": snapshot, "reversal_movement_ids": mv_ids,
    }
    if extra_detail:
        detail.update(extra_detail)
    write_audit(db, "order_void", "orders", order.id, detail,
                actor_id=actor_id, actor_name=actor_name)
    msg = f"訂單 {order.order_no} 已作廢" + (
        "（已有出貨紀錄：庫存不回補）" if shipped else "（庫存已回補）")
    if old_pay == "paid":
        msg += "；付款狀態改為已退款"
    return True, msg


@orders_bp.route("/<int:order_id>/void", methods=["POST"])
@role_required(*WRITE_ROLES)
def void(order_id):
    """作廢（必填 reason）。未出貨：staff + owner；已出貨：owner only（其餘 403）。"""
    db = get_session()
    order = db.get(Order, order_id)
    if not order:
        abort(404)
    u = current_user()
    if is_order_shipped(db, order_id) and (u is None or u.role != "owner"):
        abort(403)
    reason = (request.form.get("reason") or "").strip()
    try:
        ok, msg = perform_void(db, order, reason, _operator_name(), _uid(), _operator_name())
        if not ok:
            db.rollback()
            flash(msg)
            return redirect(url_for("orders.detail", order_id=order_id))
        db.commit()
    except Exception as exc:  # noqa: BLE001
        db.rollback()
        flash(f"作廢失敗，已全數回復：{exc}")
        return redirect(url_for("orders.detail", order_id=order_id))
    flash(msg)
    return redirect(url_for("orders.detail", order_id=order_id))


@orders_bp.route("/void-bulk", methods=["POST"])
@role_required(*WRITE_ROLES)
def void_bulk():
    """CR-6：列表勾選批次作廢。一個原因套用全批；整批同一 tx，任一失敗全數 rollback。

    - 已作廢單跳過（不重複、不算失敗）。
    - staff 勾到已出貨單 → 整批拒絕（明列單號），不部分成功；owner 可全批。
    - 不存在的 id → 整批拒絕。
    - 每筆各寫一筆 AuditLog(order_void)，detail 註 bulk=True / bulk_group_id（同批同一 uuid）。
    完成後 flash「已作廢 N 筆（回補 M 筆庫存）」並回列表（保留原篩選條件）。
    """
    db = get_session()
    back = redirect(url_for("orders.index", **request.args.to_dict()))

    raw_ids = request.form.getlist("order_ids")
    ids = []
    for v in raw_ids:
        try:
            i = int(str(v).strip())
        except (TypeError, ValueError):
            flash("勾選資料有誤（訂單 id 非整數），整批未執行")
            return back
        if i not in ids:
            ids.append(i)
    if not ids:
        flash("請先勾選要作廢的訂單")
        return back

    reason = (request.form.get("reason") or "").strip()
    if not reason:
        flash("批次作廢必須填寫原因，整批未執行")
        return back

    u = current_user()
    is_owner = bool(u and u.role == "owner")

    orders = db.query(Order).filter(Order.id.in_(ids)).all()
    found = {o.id: o for o in orders}
    missing = [i for i in ids if i not in found]
    if missing:
        flash("以下訂單不存在，整批未執行：" + "、".join(str(i) for i in missing))
        return back

    targets, skipped = [], []
    for i in ids:
        o = found[i]
        (skipped if o.voided_at is not None else targets).append(o)
    if not targets:
        flash("勾選的訂單皆已作廢，無需處理")
        return back

    shipped_map = {o.id: is_order_shipped(db, o.id) for o in targets}
    if not is_owner:
        blocked = [o.order_no for o in targets if shipped_map[o.id]]
        if blocked:
            flash("以下訂單已出貨（有出貨紀錄），作廢需由老闆（owner）操作，整批未執行："
                  + "、".join(blocked))
            return back

    group_id = uuid.uuid4().hex
    operator = _operator_name()
    actor_id = _uid()
    done = reversed_n = refunded_n = 0
    try:
        for o in targets:
            was_paid = (o.payment_status == "paid")
            ok, msg = perform_void(
                db, o, reason, operator, actor_id, operator,
                extra_detail={"bulk": True, "bulk_group_id": group_id},
            )
            if not ok:
                db.rollback()
                flash(f"訂單 {o.order_no} 作廢失敗，整批已全數回復：{msg}")
                return back
            done += 1
            if not shipped_map[o.id]:
                reversed_n += 1
            if was_paid:
                refunded_n += 1
        db.commit()
    except Exception as exc:  # noqa: BLE001
        db.rollback()
        flash(f"批次作廢失敗，已全數回復：{exc}")
        return back

    msg = f"已作廢 {done} 筆（回補 {reversed_n} 筆庫存）"
    if done - reversed_n:
        msg += f"；已出貨單 {done - reversed_n} 筆不回補"
    if refunded_n:
        msg += f"；付款狀態改為已退款 {refunded_n} 筆"
    if skipped:
        msg += "；已作廢單跳過 " + "、".join(o.order_no for o in skipped)
    flash(msg)
    return back
