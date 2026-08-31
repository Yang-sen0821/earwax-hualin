"""reports blueprint（/reports）— 報表 + 財務 + 匯出 Excel（D10 / D11 / D12 / D13）。

本檔只讀資料、做聚合與匯出，不變動任何庫存量（不碰 inventory_balances / inventory_movements 寫入）。
- D10 報表：銷售日/月/年報、商品銷售排行（依 order_items.combo_code 聚合）、庫存報表（目前/預留/已售/公關品）。
- D11 匯出 Excel：訂單 / 庫存 / 庫存異動 / 銷售報表 / 客戶排行（CR-3）五類（openpyxl）。
- D12/D13 財務：毛利 = 銷售金額 − 商品成本；淨利 = 毛利 − 包材 − 行銷 − 其他支出。
  來源：financial_entries（sale / product_cost / packaging / marketing 等分錄）+ extra_expenses（其他支出）。

權限（§六）：reports / finance 讀取 → 全角色可讀（viewer 以上）；故僅 @login_required。
"""
from datetime import datetime, timedelta
from io import BytesIO
from decimal import Decimal

from flask import (
    Blueprint, render_template, request, send_file,
    redirect, url_for, flash, abort,
)
from sqlalchemy import func, case

from auth import login_required, role_required, current_user
from db import (
    get_session,
    Order, OrderItem, Product, SalesPlan, Customer,
    InventoryBalance, InventoryMovement,
    FinancialEntry, ExtraExpense, Setting,
    COMBO_CODES, active_orders,
)
from display_labels import (
    combo_label, pool_label, payment_label, shipping_label, mtype_label,
    shipping_method_label, action_label, target_type_label,
)
from audit_util import write_audit, snapshot, diff, summarize
import inventory_service as inv

reports_bp = Blueprint("reports", __name__, url_prefix="/reports")

# CR-8 支出留痕欄位
EXPENSE_FIELDS = ("name", "amount", "category", "expense_date", "note")

# =========================================================================
# 支出管理：唯一成本來源 = extra_expenses by category
# 類別固定下拉（單一成本來源；ROI/財務一律依此加總，禁止第二套來源）
# =========================================================================
EXPENSE_CATEGORIES = ("設備", "面膜進貨成本", "建檔費", "包材", "行銷", "其他")

# ROI/財務成本側對照：類別 → 聚合鍵
CATEGORY_PRODUCT = "面膜進貨成本"   # 商品成本
CATEGORY_PACKAGING = "包材"         # 包材成本
CATEGORY_MARKETING = "行銷"         # 行銷成本
CATEGORY_SETUP = "建檔費"           # 建檔額外支出
CATEGORY_EQUIPMENT = "設備"         # 設備（投資，計入總投入與回本門檻）
CATEGORY_OTHER = "其他"             # 其他支出


def _uid():
    u = current_user()
    return u.id if u else None


# =========================================================================
# 共用：combo_code 顯示名稱（依 sales_plans seed；查不到退回 code 本身）
# =========================================================================
def _combo_label_map(db):
    rows = db.query(SalesPlan.combo_code, SalesPlan.name).all()
    m = {code: name for code, name in rows}
    # 新模型單位（建單只產生這兩個）：LOOSE=片、BOX=盒
    m.setdefault("LOOSE", "片")
    m.setdefault("BOX", "盒")
    # 舊四個閉集 code 相容（歷史資料）
    fallback = {"SINGLE": "單片體驗組", "BOX1": "經典盒裝",
                "BOX3": "植萃養膚組", "BOX10": "尊寵囤貨組"}
    for c in COMBO_CODES:
        m.setdefault(c, fallback.get(c, c))
    return m


def _to_float(v):
    if v is None:
        return 0.0
    if isinstance(v, Decimal):
        return float(v)
    return float(v)


# =========================================================================
# 共用：銷售期間分組標籤（SQLite / Postgres 皆相容，於 Python 端分組）
# =========================================================================
def _period_key(dt, granularity):
    if dt is None:
        return "(未知)"
    if granularity == "day":
        return dt.strftime("%Y-%m-%d")
    if granularity == "year":
        return dt.strftime("%Y")
    return dt.strftime("%Y-%m")  # month（預設）


def _sales_by_period(db, granularity="month"):
    """以 orders 為基礎，依期間聚合訂單數與銷售金額。

    銷售金額採 orders.total_amount（下單當下總額快照）。
    """
    # CR-4：作廢單排除於統計
    rows = active_orders(db.query(Order.created_at, Order.total_amount, Order.shipping_fee)).all()
    bucket = {}
    for created_at, total, fee in rows:
        key = _period_key(created_at, granularity)
        b = bucket.setdefault(key, {"period": key, "order_count": 0,
                                    "amount": 0.0, "shipping_fee": 0.0})
        b["order_count"] += 1
        b["amount"] += _to_float(total)
        b["shipping_fee"] += _to_float(fee)   # CR-5：代收運費另列，不混入銷售金額
    return sorted(bucket.values(), key=lambda r: r["period"])


# =========================================================================
# 共用：商品銷售排行（依 order_items.combo_code 聚合 單片/盒裝/3盒/10盒）
# =========================================================================
def _sales_ranking(db):
    label = _combo_label_map(db)
    # CR-4：join orders 排除作廢單
    rows = (
        active_orders(
            db.query(
                OrderItem.combo_code,
                func.count(OrderItem.id),
                func.coalesce(func.sum(OrderItem.qty), 0),
                func.coalesce(func.sum(OrderItem.subtotal), 0),
            )
            .join(Order, Order.id == OrderItem.order_id)
        )
        .group_by(OrderItem.combo_code)
        .all()
    )
    result = []
    for combo_code, line_count, total_qty, total_amount in rows:
        result.append({
            "combo_code": combo_code,
            "combo_name": label.get(combo_code, combo_code),
            "line_count": int(line_count or 0),
            "total_qty": int(total_qty or 0),
            "total_amount": _to_float(total_amount),
        })
    result.sort(key=lambda r: r["total_qty"], reverse=True)
    return result


# =========================================================================
# 共用：客戶銷售排行（CR-3，依 orders.customer_id 聚合）
# =========================================================================
UNBOUND_CUSTOMER_LABEL = "未綁定客戶"


def _customer_ranking(db, date_from=None, date_to=None):
    """依客戶聚合訂單（口徑與 customers.customer_summary 一致）：
      訂單數     = 非作廢單數（含已退款單）
      累計金額   = 非作廢 且 payment_status != refunded 的 total_amount 合計（不含運費）
      最近購買   = 最近一筆非作廢單 created_at
    date_from / date_to（datetime，皆含當日）依 created_at 篩選；None = 不限。
    回傳 list[dict(customer_id, name, phone, order_count, total_amount, last_order_at)]，
    金額 desc、次鍵訂單數 desc；customer_id IS NULL 的單另計一列「未綁定客戶」固定排最後。
    """
    amount_expr = case(
        (Order.payment_status != "refunded", Order.total_amount), else_=0
    )
    q = active_orders(
        db.query(
            Order.customer_id,
            func.count(Order.id),
            func.coalesce(func.sum(amount_expr), 0),
            func.max(Order.created_at),
        )
    )
    if date_from is not None:
        q = q.filter(Order.created_at >= date_from)
    if date_to is not None:
        # 含 date_to 當日：< 次日 00:00
        q = q.filter(Order.created_at < (date_to.replace(hour=0, minute=0, second=0, microsecond=0)
                                         + timedelta(days=1)))
    rows = q.group_by(Order.customer_id).all()
    cust = {c.id: c for c in db.query(Customer).all()}
    ranked, unbound = [], None
    for cid, cnt, amt, last in rows:
        row = {
            "customer_id": cid,
            "name": cust[cid].name if cid in cust else (UNBOUND_CUSTOMER_LABEL if cid is None else f"#{cid}"),
            "phone": (cust[cid].phone or "") if cid in cust else "",
            "order_count": int(cnt or 0),
            "total_amount": _to_float(amt),
            "last_order_at": last,
        }
        if cid is None:
            unbound = row
        else:
            ranked.append(row)
    ranked.sort(key=lambda r: (-r["total_amount"], -r["order_count"]))
    if unbound:
        ranked.append(unbound)
    return ranked


# =========================================================================
# 共用：庫存報表（目前 / 預留 / 已售 / 公關品）
# =========================================================================
def _inventory_report(db):
    """彙整各 (商品 × 池) 的庫存切片：
      - 目前(normal) / 預留(reserved) / 公關品(pr) / 試用(trial) / 損耗(scrap)
      - 已售：由 inventory_movements SALE（負）+ SALE_REVERSAL（正，CR-4 回補）淨額取絕對值
    """
    # balances：依 product 聚合各 category
    bal_rows = (
        db.query(
            InventoryBalance.product_id,
            InventoryBalance.inventory_pool,
            InventoryBalance.stock_category,
            InventoryBalance.qty,
            InventoryBalance.unit,
        )
        .all()
    )
    # 已售：SALE（負）+ SALE_REVERSAL（正）淨額；編輯/作廢回補後已售自動扣回（CR-4）
    sold_rows = (
        db.query(
            InventoryMovement.product_id,
            func.coalesce(func.sum(InventoryMovement.qty_delta), 0),
        )
        .filter(InventoryMovement.movement_type.in_(("SALE", "SALE_REVERSAL")))
        .group_by(InventoryMovement.product_id)
        .all()
    )
    sold_map = {pid: max(0, -int(s or 0)) for pid, s in sold_rows}

    # 商品名稱
    prods = {p.id: p for p in db.query(Product).all()}

    # 依 (product_id, pool) 組合彙整
    grid = {}
    for pid, pool, cat, qty, unit in bal_rows:
        key = (pid, pool)
        row = grid.setdefault(key, {
            "product_id": pid,
            "product_name": prods[pid].name if pid in prods else f"#{pid}",
            "sku": prods[pid].sku if pid in prods else "",
            "inventory_pool": pool,
            "unit": unit,
            "normal": 0, "reserved": 0, "pr": 0, "trial": 0, "scrap": 0,
        })
        if cat in row:
            row[cat] += int(qty or 0)

    result = list(grid.values())
    # CR-9：歷史輸入 = 合計 + 累計消耗（該池 SALE 淨額 + 出庫類 + 紙袋出貨；校正／轉移不計）
    consumed_map = inv.consumed_by_pool(db)
    # 已售掛在商品層（不分池）→ 標在該商品第一個池列，其餘填 0 避免重複計
    counted = set()
    for row in result:
        # 合計 = 五分類加總（2026-08-31：森哥只看「目前」欄誤以為庫存未修正）
        row["total"] = sum(row[c] for c in ("normal", "reserved", "pr", "trial", "scrap"))
        row["consumed"] = consumed_map.get((row["product_id"], row["inventory_pool"]), 0)
        row["hist_in"] = row["total"] + row["consumed"]
        pid = row["product_id"]
        if pid not in counted:
            row["sold"] = sold_map.get(pid, 0)
            counted.add(pid)
        else:
            row["sold"] = 0
    result.sort(key=lambda r: (r["product_id"], r["inventory_pool"]))
    return result


# =========================================================================
# 共用：成本聚合（唯一來源）= extra_expenses 依 category 加總
# ROI 與財務的成本側全部走這裡，絕不再從 settings cost_*_test 或
# financial_entries(product_cost/packaging/marketing) 取成本，避免重複計算。
# =========================================================================
def _cost_by_category(db):
    """回傳各成本類別小計（dict）+ 衍生的彙整欄位。

    商品成本 = Σ(面膜進貨成本)
    包材成本 = Σ(包材)
    行銷成本 = Σ(行銷)
    建檔額外支出 = Σ(建檔費)
    設備 = Σ(設備)  ← 投資，計入總投入成本與回本門檻
    其他支出 = Σ(其他)
    總投入成本 = 以上全部加總
    """
    rows = (
        db.query(
            ExtraExpense.category,
            func.coalesce(func.sum(ExtraExpense.amount), 0),
        )
        .group_by(ExtraExpense.category)
        .all()
    )
    by_cat = {cat: 0.0 for cat in EXPENSE_CATEGORIES}
    uncategorized = 0.0
    for cat, amt in rows:
        a = _to_float(amt)
        if cat in by_cat:
            by_cat[cat] += a
        else:
            # 類別空白/非閉集者，歸入「其他」以免漏算（單一來源不丟資料）
            uncategorized += a
    by_cat[CATEGORY_OTHER] += uncategorized

    product_cost = by_cat[CATEGORY_PRODUCT]
    packaging_cost = by_cat[CATEGORY_PACKAGING]
    marketing_cost = by_cat[CATEGORY_MARKETING]
    setup_cost = by_cat[CATEGORY_SETUP]
    equipment_cost = by_cat[CATEGORY_EQUIPMENT]
    other_cost = by_cat[CATEGORY_OTHER]
    total_cost = (product_cost + packaging_cost + marketing_cost
                  + setup_cost + equipment_cost + other_cost)
    return {
        "by_cat": by_cat,
        "product_cost": product_cost,
        "packaging_cost": packaging_cost,
        "marketing_cost": marketing_cost,
        "setup_cost": setup_cost,
        "equipment_cost": equipment_cost,
        "other_cost": other_cost,
        "total_cost": total_cost,
    }


# =========================================================================
# 共用：財務報表（毛利 / 淨利）D12 / D13
# 成本來源：唯一 = extra_expenses by category（不再讀 financial_entries 成本分錄）
# 收入來源：financial_entries(sale) 優先，無則回退 orders.total_amount（維持不動）
# =========================================================================
def _financial_report(db):
    """毛利 = 銷售金額 − 商品成本
       淨利 = 銷售金額 − 總投入成本

    成本側單一來源 = extra_expenses by category：
      商品成本=面膜進貨成本、包材=包材、行銷=行銷、
      建檔=建檔費、設備=設備、其他=其他 → 總投入成本為全部加總。
    收入側維持原口徑（financial_entries sale 或 orders 總額）。
    """
    cost = _cost_by_category(db)

    # ---- 收入側（不動）----
    sale_entries = _to_float(
        db.query(func.coalesce(func.sum(FinancialEntry.amount), 0))
        .filter(FinancialEntry.entry_type == "sale")
        .scalar()
    )
    orders_total = _to_float(
        active_orders(db.query(func.coalesce(func.sum(Order.total_amount), 0))).scalar()
    )
    sales_amount = sale_entries if sale_entries > 0 else orders_total

    product_cost = cost["product_cost"]
    total_cost = cost["total_cost"]

    gross_profit = sales_amount - product_cost
    net_profit = sales_amount - total_cost

    return {
        "sales_amount": sales_amount,
        "orders_total": orders_total,
        "product_cost": product_cost,
        "packaging_cost": cost["packaging_cost"],
        "marketing_cost": cost["marketing_cost"],
        "setup_cost": cost["setup_cost"],
        "equipment_cost": cost["equipment_cost"],
        "other_expense": cost["other_cost"],
        "total_cost": total_cost,
        "gross_profit": gross_profit,
        "net_profit": net_profit,
        "by_cat": cost["by_cat"],
    }


# =========================================================================
# 共用：成本與利潤（投報率 ROI）— 限 owner / accounting
# =========================================================================
def _roi_report(db):
    """成本與利潤（投報率 ROI）聚合，唯讀。

    累積總收入口徑（不動）：
      - 優先採 financial_entries 的 sale 分錄加總；
      - 若無 sale 分錄，則以「已認列營收」訂單為準＝orders 中
        payment_status='paid' 或 shipping_status ∈ (shipped, delivered) 的 total_amount 加總。

    成本各項口徑（唯一來源 = extra_expenses by category）：
      - 商品成本     = Σ(面膜進貨成本)
      - 包材成本     = Σ(包材)
      - 行銷成本     = Σ(行銷)
      - 建檔額外支出 = Σ(建檔費)
      - 設備         = Σ(設備)  ← 投資，計入總投入成本與回本門檻
      - 其他支出     = Σ(其他)
      總投入成本 ＝ 以上全部加總

    淨利 ＝ 總收入 − 總投入成本
    毛利 ＝ 總收入 − 商品成本
    ROI% ＝ 淨利 ÷ 總投入成本 × 100（總成本為 0 時回 None → 前台顯示 N/A，不除以零）
    """
    # ---- 累積總收入（收入側維持不動）----
    sale_entries = _to_float(
        db.query(func.coalesce(func.sum(FinancialEntry.amount), 0))
        .filter(FinancialEntry.entry_type == "sale")
        .scalar()
    )
    recognized_orders_total = _to_float(
        active_orders(db.query(func.coalesce(func.sum(Order.total_amount), 0)))
        .filter(
            (Order.payment_status == "paid")
            | (Order.shipping_status.in_(("shipped", "delivered")))
        )
        .scalar()
    )
    if sale_entries > 0:
        total_revenue = sale_entries
        revenue_basis = "financial_entries(sale)"
    else:
        total_revenue = recognized_orders_total
        revenue_basis = "orders(已付款/已出貨)"

    # ---- 成本各項（唯一來源 = extra_expenses by category）----
    cost = _cost_by_category(db)
    product_cost = cost["product_cost"]
    packaging_cost = cost["packaging_cost"]
    marketing_cost = cost["marketing_cost"]
    setup_cost = cost["setup_cost"]
    equipment_cost = cost["equipment_cost"]
    other_expense = cost["other_cost"]
    total_cost = cost["total_cost"]

    gross_profit = total_revenue - product_cost
    net_profit = total_revenue - total_cost
    roi_pct = (net_profit / total_cost * 100) if total_cost > 0 else None

    # ---- 回本計算（沿用既有「總投入成本」與「累積收入」口徑，不重算）----
    # 回本門檻 = 總投入成本；已回收 = 累積收入
    # 回本進度% = 累積收入 ÷ 總投入成本 × 100（總成本 0 → None → 前台 N/A，防除零）
    # 尚需回收 = max(0, 總投入成本 − 累積收入)
    # 累積收入 ≥ 總投入成本 → 已回本，淨賺 = 累積收入 − 總投入成本
    payback_threshold = total_cost          # 回本門檻
    recovered = total_revenue               # 已回收
    payback_pct = (recovered / total_cost * 100) if total_cost > 0 else None
    remaining_to_recover = max(0.0, total_cost - total_revenue)
    paid_back = (total_cost > 0) and (total_revenue >= total_cost)
    net_earned = (total_revenue - total_cost) if paid_back else 0.0

    return {
        "total_revenue": total_revenue,
        "revenue_basis": revenue_basis,
        "sale_entries": sale_entries,
        "recognized_orders_total": recognized_orders_total,
        "product_cost": product_cost,
        "packaging_cost": packaging_cost,
        "marketing_cost": marketing_cost,
        "setup_cost": setup_cost,
        "equipment_cost": equipment_cost,
        "other_expense": other_expense,
        "by_cat": cost["by_cat"],
        "total_cost": total_cost,
        "gross_profit": gross_profit,
        "net_profit": net_profit,
        "roi_pct": roi_pct,
        # ---- 回本計算 ----
        "payback_threshold": payback_threshold,
        "recovered": recovered,
        "payback_pct": payback_pct,
        "remaining_to_recover": remaining_to_recover,
        "paid_back": paid_back,
        "net_earned": net_earned,
    }


# =========================================================================
# 頁面：報表首頁（彙整入口 + 簡要 KPI）
# =========================================================================
@reports_bp.route("/")
@role_required("owner", "accounting")  # 2026-07-12 收緊：金額/毛利相關，staff 不可見
def index():
    db = get_session()
    fin = _financial_report(db)
    ranking = _sales_ranking(db)
    kpi = {
        "order_count": active_orders(db.query(func.count(Order.id))).scalar() or 0,
        "customer_count": db.query(func.count(Customer.id)).scalar() or 0,
        "sales_amount": fin["sales_amount"],
        "net_profit": fin["net_profit"],
    }
    return render_template("reports/index.html", section="reports",
                           kpi=kpi, ranking=ranking, fin=fin)


# =========================================================================
# 頁面：銷售報表（日 / 月 / 年）+ 排行
# =========================================================================
@reports_bp.route("/sales")
@role_required("owner", "accounting")  # 2026-07-12 收緊：金額/毛利相關，staff 不可見
def sales():
    db = get_session()
    granularity = request.args.get("g", "month")
    if granularity not in ("day", "month", "year"):
        granularity = "month"
    periods = _sales_by_period(db, granularity)
    ranking = _sales_ranking(db)
    totals = {
        "order_count": sum(p["order_count"] for p in periods),
        "amount": sum(p["amount"] for p in periods),
        "shipping_fee": sum(p["shipping_fee"] for p in periods),
    }
    return render_template("reports/sales.html", section="reports",
                           granularity=granularity, periods=periods,
                           ranking=ranking, totals=totals)


# =========================================================================
# 頁面：客戶銷售排行（CR-3；owner / accounting，與其他金額報表同守門）
# =========================================================================
@reports_bp.route("/customers")
@role_required("owner", "accounting")
def customers():
    db = get_session()
    ok_f, date_from, err_f = _parse_date(request.args.get("from"))
    ok_t, date_to, err_t = _parse_date(request.args.get("to"))
    error = None
    if not ok_f or not ok_t:
        error = err_f or err_t
        date_from = date_to = None
    ranking = _customer_ranking(db, date_from, date_to)
    bound = [r for r in ranking if r["customer_id"] is not None]
    totals = {
        "order_count": sum(r["order_count"] for r in ranking),
        "amount": sum(r["total_amount"] for r in ranking),
        "customer_count": len(bound),
    }
    return render_template(
        "reports/customers.html", section="customer_ranking",
        ranking=ranking, totals=totals, error=error,
        f_from=(request.args.get("from") or "").strip() if ok_f else "",
        f_to=(request.args.get("to") or "").strip() if ok_t else "",
    )


# =========================================================================
# 頁面：庫存報表（目前 / 預留 / 已售 / 公關品）
# =========================================================================
@reports_bp.route("/inventory")
@login_required
def inventory():
    db = get_session()
    rows = _inventory_report(db)
    return render_template("reports/inventory.html", section="reports", rows=rows)


# =========================================================================
# 頁面：財務報表（毛利 / 淨利）
# =========================================================================
@reports_bp.route("/finance")
@role_required("owner", "accounting")  # 2026-07-12 收緊：金額/毛利相關，staff 不可見
def finance():
    db = get_session()
    fin = _financial_report(db)
    return render_template("reports/finance.html", section="reports", fin=fin)


# =========================================================================
# 頁面：成本與利潤（投報率 ROI）— 財務機密，限 owner / accounting
# =========================================================================
@reports_bp.route("/roi")
@role_required("owner", "accounting")
def roi():
    """ROI 頁：唯讀。成本各類別明細只顯示，不在此編輯（編輯走支出管理頁）。"""
    db = get_session()
    roi_data = _roi_report(db)
    return render_template(
        "reports/roi.html", section="reports", roi=roi_data,
        categories=EXPENSE_CATEGORIES,
    )


# =========================================================================
# 支出管理（/reports/expenses）— 唯一成本來源：extra_expenses by category
# 品項列管：name + amount + category + expense_date + note（vendor 不在 schema，置於 note）
# 權限：owner / accounting（@role_required）
# 收入、已售、庫存維持唯讀自動，本模組不碰。
# =========================================================================
def _parse_amount(raw):
    """解析金額字串為 float；回 (ok, value, err)。空/非數值/負數回錯。"""
    raw = (raw or "").strip()
    try:
        v = float(raw)
    except (TypeError, ValueError):
        return False, None, "金額需為數字"
    if v < 0:
        return False, None, "金額不得為負"
    return True, v, None


def _parse_date(raw):
    """解析 YYYY-MM-DD；空字串回 (True, None)；格式錯回 (False, None, err)。"""
    raw = (raw or "").strip()
    if not raw:
        return True, None, None
    try:
        return True, datetime.strptime(raw, "%Y-%m-%d"), None
    except ValueError:
        return False, None, "日期格式需為 YYYY-MM-DD"


@reports_bp.route("/expenses")
@role_required("owner", "accounting")
def expenses():
    db = get_session()
    rows = db.query(ExtraExpense).order_by(
        ExtraExpense.category, ExtraExpense.id).all()
    # 依類別分組（依固定下拉順序），各類別小計 + 總計
    grouped = {cat: [] for cat in EXPENSE_CATEGORIES}
    other_key = CATEGORY_OTHER
    subtotals = {cat: 0.0 for cat in EXPENSE_CATEGORIES}
    grand_total = 0.0
    for e in rows:
        cat = e.category if e.category in grouped else other_key
        grouped[cat].append(e)
        amt = _to_float(e.amount)
        subtotals[cat] += amt
        grand_total += amt
    groups = [
        {"category": cat, "rows": grouped[cat], "subtotal": subtotals[cat]}
        for cat in EXPENSE_CATEGORIES
    ]
    return render_template(
        "reports/expenses.html", section="expenses",
        groups=groups, grand_total=grand_total,
        categories=EXPENSE_CATEGORIES,
    )


@reports_bp.route("/expenses/add", methods=["POST"])
@role_required("owner", "accounting")
def expense_add():
    db = get_session()
    name = request.form.get("name", "").strip()
    category = request.form.get("category", "").strip()
    if not name:
        flash("品項名稱為必填", "error")
        return redirect(url_for("reports.expenses"))
    if category not in EXPENSE_CATEGORIES:
        flash("類別不合法（須為固定下拉之一）", "error")
        return redirect(url_for("reports.expenses"))
    ok, val, err = _parse_amount(request.form.get("amount", ""))
    if not ok:
        flash(f"支出新增失敗：{err}", "error")
        return redirect(url_for("reports.expenses"))
    dok, dt, derr = _parse_date(request.form.get("expense_date", ""))
    if not dok:
        flash(f"支出新增失敗：{derr}", "error")
        return redirect(url_for("reports.expenses"))
    e = ExtraExpense(
        name=name, amount=val, category=category,
        expense_date=dt or datetime.utcnow(),
        note=request.form.get("note", "").strip() or None,
        created_by=_uid(),
    )
    db.add(e)
    db.flush()
    write_audit(db, "expense_create", "extra_expenses", e.id,
                {"after": snapshot(e, EXPENSE_FIELDS)})   # CR-8
    db.commit()
    flash(f"支出「{name}」已新增（{category} {val:.2f}）", "ok")
    return redirect(url_for("reports.expenses"))


@reports_bp.route("/expenses/<int:eid>/update", methods=["POST"])
@role_required("owner", "accounting")
def expense_update(eid):
    db = get_session()
    e = db.query(ExtraExpense).get(eid)
    if not e:
        abort(404)
    name = request.form.get("name", "").strip()
    category = request.form.get("category", "").strip()
    if not name:
        flash("品項名稱為必填", "error")
        return redirect(url_for("reports.expenses"))
    if category not in EXPENSE_CATEGORIES:
        flash("類別不合法（須為固定下拉之一）", "error")
        return redirect(url_for("reports.expenses"))
    ok, val, err = _parse_amount(request.form.get("amount", ""))
    if not ok:
        flash(f"支出更新失敗：{err}", "error")
        return redirect(url_for("reports.expenses"))
    dok, dt, derr = _parse_date(request.form.get("expense_date", ""))
    if not dok:
        flash(f"支出更新失敗：{derr}", "error")
        return redirect(url_for("reports.expenses"))
    before = snapshot(e, EXPENSE_FIELDS)
    e.name = name
    e.amount = val
    e.category = category
    if dt is not None:
        e.expense_date = dt
    e.note = request.form.get("note", "").strip() or None
    b, a = diff(before, snapshot(e, EXPENSE_FIELDS))
    if a:   # CR-8：只寫有變的欄
        write_audit(db, "expense_update", "extra_expenses", e.id,
                    {"name": e.name, "before": b, "after": a})
    db.commit()
    flash(f"支出「{name}」已更新（{category} {val:.2f}）", "ok")
    return redirect(url_for("reports.expenses"))


@reports_bp.route("/expenses/<int:eid>/delete", methods=["POST"])
@role_required("owner", "accounting")
def expense_delete(eid):
    db = get_session()
    e = db.query(ExtraExpense).get(eid)
    if not e:
        abort(404)
    nm = e.name
    write_audit(db, "expense_delete", "extra_expenses", e.id,
                {"before": snapshot(e, EXPENSE_FIELDS)})   # CR-8
    db.delete(e)
    db.commit()
    flash(f"支出「{nm}」已刪除", "ok")
    return redirect(url_for("reports.expenses"))


# =========================================================================
# 匯出 Excel（D11）— 單一按鈕、一份 .xlsx、多工作表分頁。
# 分頁：訂單 / 庫存 / 庫存異動 / 銷售報表（期間 + 排行）。
# 內容全中文（沿用 display_labels 對照：狀態 / 組合 / 庫存池 / 異動類別）。
# 權限：@login_required（沿用原匯出頁角色限制：viewer 以上可讀可匯）。
# =========================================================================
def _new_workbook():
    from openpyxl import Workbook
    from openpyxl.styles import Font
    wb = Workbook()
    return wb, Font


def _write_header(ws, headers, Font):
    ws.append(headers)
    bold = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold


def _send_xlsx(wb, filename):
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _ts():
    return datetime.utcnow().strftime("%Y%m%d_%H%M%S")


def _sheet_orders(wb, db, Font):
    """工作表：訂單（首頁，沿用 wb.active，避免留空白 Sheet）。"""
    ws = wb.active
    ws.title = "訂單"
    # CR-5：總金額=商品實收−折扣（不含運費）；另列 折扣 / 運費 / 運送方式 / 應收合計
    _write_header(ws, [
        "訂單編號", "客戶", "收件人", "電話", "地址",
        "總金額", "折扣", "運費", "應收合計", "運送方式", "運送備註",
        "付款狀態", "出貨狀態", "備註", "建立時間",
    ], Font)
    cust = {c.id: c.name for c in db.query(Customer).all()}
    for o in active_orders(db.query(Order)).order_by(Order.id).all():   # CR-4：作廢單不匯出
        total = _to_float(o.total_amount)
        fee = _to_float(o.shipping_fee)
        ws.append([
            o.order_no,
            cust.get(o.customer_id, ""),
            o.recipient_name or "",
            o.recipient_phone or "",
            o.shipping_address or "",
            total,
            _to_float(o.discount),
            fee,
            total + fee,
            shipping_method_label(o.shipping_method),
            o.shipping_note or "",
            payment_label(o.payment_status),
            shipping_label(o.shipping_status),
            o.note or "",
            o.created_at.strftime("%Y-%m-%d %H:%M") if o.created_at else "",
        ])


def _sheet_inventory(wb, db, Font):
    """工作表：庫存。"""
    ws = wb.create_sheet("庫存")
    _write_header(ws, [
        "SKU", "商品", "庫存池", "目前", "預留",
        "公關品", "試用品", "損耗報廢", "已售(累計)", "單位",
    ], Font)
    for r in _inventory_report(db):
        ws.append([
            r["sku"], r["product_name"], pool_label(r["inventory_pool"]),
            r["normal"], r["reserved"], r["pr"], r["trial"], r["scrap"],
            r["sold"], r["unit"] or "",
        ])


def _sheet_movements(wb, db, Font):
    """工作表：庫存異動。"""
    ws = wb.create_sheet("庫存異動")
    _write_header(ws, [
        "異動ID", "SKU", "商品", "庫存池", "從分類", "到分類",
        "變動量", "變動前", "變動後", "異動類別", "來源類型", "來源ID",
        "群組ID", "原因", "經手人", "備註", "時間",
    ], Font)
    prods = {p.id: p for p in db.query(Product).all()}
    for m in db.query(InventoryMovement).order_by(InventoryMovement.movement_id).all():
        p = prods.get(m.product_id)
        ws.append([
            m.movement_id,
            p.sku if p else "",
            p.name if p else f"#{m.product_id}",
            pool_label(m.inventory_pool),
            m.stock_category_from or "",
            m.stock_category_to or "",
            m.qty_delta,
            m.qty_before,
            m.qty_after,
            mtype_label(m.movement_type),
            m.ref_type or "",
            m.ref_id or "",
            m.movement_group_id or "",
            m.reason or "",
            m.operator or "",
            m.note or "",
            m.created_at.strftime("%Y-%m-%d %H:%M") if m.created_at else "",
        ])


def _sheet_sales(wb, db, Font, granularity):
    """工作表：銷售報表（期間銷售 + 商品銷售排行，合併同一分頁）。"""
    ws = wb.create_sheet("銷售報表")
    label = {"day": "日報", "month": "月報", "year": "年報"}[granularity]

    # 區塊一：期間銷售
    ws.append([f"期間銷售（{label}）"])
    ws["A1"].font = Font(bold=True)
    _write_header(ws, [f"期間（{label}）", "訂單數", "銷售金額", "運費", "應收合計"], Font)
    for p in _sales_by_period(db, granularity):
        ws.append([p["period"], p["order_count"], p["amount"],
                   p["shipping_fee"], p["amount"] + p["shipping_fee"]])

    # 空一列後接區塊二：商品銷售排行
    ws.append([])
    title_row = ws.max_row + 1
    ws.append(["商品銷售排行"])
    ws[f"A{title_row}"].font = Font(bold=True)
    _write_header(ws, ["組合代碼", "組合名稱", "明細筆數", "銷售組數", "銷售金額"], Font)
    for r in _sales_ranking(db):
        ws.append([r["combo_code"], combo_label(r["combo_code"]),
                   r["line_count"], r["total_qty"], r["total_amount"]])


def _sheet_customers(wb, db, Font):
    """工作表：客戶排行（CR-3；全期，未綁定客戶另列最後、不編名次）。"""
    ws = wb.create_sheet("客戶排行")
    _write_header(ws, ["名次", "客戶", "電話", "訂單數", "累計金額", "最近購買"], Font)
    rank = 0
    for r in _customer_ranking(db):
        if r["customer_id"] is not None:
            rank += 1
        ws.append([
            rank if r["customer_id"] is not None else "",
            r["name"], r["phone"], r["order_count"], r["total_amount"],
            r["last_order_at"].strftime("%Y-%m-%d") if r["last_order_at"] else "",
        ])


AUDIT_EXPORT_LIMIT = 1000


def _sheet_audit(wb, db, Font):
    """工作表：操作紀錄（CR-8；最近 1000 筆，時間降冪，台北時間；摘要沿用 audit_util.summarize）。"""
    from db import AuditLog, User
    from audit_util import fmt_ts, parse_detail
    ws = wb.create_sheet("操作紀錄")
    _write_header(ws, ["時間", "帳號", "動作", "對象類型", "對象ID", "摘要", "詳細(JSON)"], Font)
    users = {u.id: u for u in db.query(User).all()}
    logs = (db.query(AuditLog).order_by(AuditLog.created_at.desc(), AuditLog.id.desc())
            .limit(AUDIT_EXPORT_LIMIT).all())
    for a in logs:
        u = users.get(a.actor_id) if a.actor_id else None
        actor = a.actor_name or (u.display_name if u else None) or (u.username if u else None) or "system"
        if u and u.username and u.username != actor:
            actor = f"{actor}（{u.username}）"
        d = parse_detail(a.detail)
        ws.append([
            fmt_ts(a.created_at), actor, action_label(a.action),
            target_type_label(a.target_type), a.target_id or "",
            summarize(a.action, d), (a.detail or "")[:32000],
        ])


@reports_bp.route("/export.xlsx")
@role_required("owner", "accounting")  # 2026-07-12 收緊：金額/毛利相關，staff 不可見
def export_all():
    """單鍵匯出：一份 .xlsx，含訂單 / 庫存 / 庫存異動 / 銷售報表 / 客戶排行 / 操作紀錄 六個工作表。"""
    db = get_session()
    granularity = request.args.get("g", "month")
    if granularity not in ("day", "month", "year"):
        granularity = "month"

    wb, Font = _new_workbook()
    _sheet_orders(wb, db, Font)       # 首頁，用 wb.active（不留空白 Sheet）
    _sheet_inventory(wb, db, Font)
    _sheet_movements(wb, db, Font)
    _sheet_sales(wb, db, Font, granularity)
    _sheet_customers(wb, db, Font)    # CR-3
    _sheet_audit(wb, db, Font)        # CR-8

    return _send_xlsx(wb, f"flora_court_{_ts()}.xlsx")
