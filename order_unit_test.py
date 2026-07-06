"""throwaway 驗收：建單品項簡化為 片/盒 + 數量自填，扣減直接對應裸片/盒裝池（R1 不變）。

不污染交付庫：DATABASE_URL 指向 %TEMP% 的全新 sqlite。
驗證項目（對照接案需求）：
  ① 建單頁(桌面 /orders/new + 手機 /m/orders/new) 品項下拉只剩 片/盒 兩選項
  ② 盒×3 → 盒裝 normal 池 -3、寫 SALE movement、金額正確；片×5 → 裸片池 -5
  ③ 多列訂單(盒×2 + 片×4) 同交易扣兩池；任一缺貨整張 rollback（不扣不留痕）
  ④ 扣不到 reserved（預留池不動）
  ⑤ 訂單明細 / 銷售排行 / Excel 訂單分頁顯示「片/盒」中文
  ⑥ ROI / 支出 / 庫存 / 權限其他功能沒壞
"""
import os
import sys
import tempfile

# 強制 UTF-8 輸出（Windows cp950 主控台會炸中文/特殊符號）
try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except Exception:
    pass

# ---- 指向 throwaway sqlite（import config/app 之前必須先設好）----
_TMP_DB = os.path.join(tempfile.gettempdir(), "flora_court_unit_test.db")
if os.path.exists(_TMP_DB):
    os.remove(_TMP_DB)
os.environ["DATABASE_URL"] = "sqlite:///" + _TMP_DB.replace("\\", "/")
os.environ["ENABLE_MOBILE"] = "true"

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import init_db
from db import (
    get_session, Order, OrderItem, InventoryBalance, InventoryMovement, Product,
)
from app import create_app

PASS = []
FAIL = []


def check(name, cond, detail=""):
    (PASS if cond else FAIL).append(name)
    mark = "PASS" if cond else "FAIL"
    print(f"[{mark}] {name}" + (f" — {detail}" if detail else ""))


def _bal(db, pid, pool, cat="normal"):
    b = db.query(InventoryBalance).filter_by(
        product_id=pid, inventory_pool=pool, stock_category=cat).first()
    return b.qty if b else None


def login(client, username="owner", password="owner123"):
    return client.post("/login", data={"username": username, "password": password},
                       follow_redirects=True)


def main():
    init_db.create_all()
    session = get_session()
    try:
        init_db.seed(session)
    finally:
        session.close()

    app = create_app()
    app.config["WTF_CSRF_ENABLED"] = False
    client = app.test_client()
    login(client)

    db = get_session()
    mask = db.query(Product).filter_by(sku="FC-MASK-001").first()
    pid = mask.id

    box0 = _bal(db, pid, "boxed")
    loose0 = _bal(db, pid, "loose_piece")
    box_res0 = _bal(db, pid, "boxed", "reserved")
    print(f"初始：盒裝normal={box0} 裸片normal={loose0} 盒裝reserved={box_res0}")
    db.close()

    # ---- ① 建單頁只剩 片/盒 ----
    html = client.get("/orders/new").get_data(as_text=True)
    desktop_has_units = ('"LOOSE"' in html and '"BOX"' in html)
    desktop_no_old = not any(c in html for c in ('"SINGLE"', '"BOX1"', '"BOX3"', '"BOX10"'))
    check("①桌面建單頁含 LOOSE/BOX 單位", desktop_has_units)
    check("①桌面建單頁不含舊 4-combo", desktop_no_old, "" if desktop_no_old else "仍出現舊 combo")

    mhtml = client.get("/m/orders/new").get_data(as_text=True)
    mob_has_units = ('value="LOOSE"' in mhtml and 'value="BOX"' in mhtml)
    mob_no_old = not any(f'value="{c}"' in mhtml for c in ("SINGLE", "BOX1", "BOX3", "BOX10"))
    check("①手機建單頁含 片/盒 單位", mob_has_units)
    check("①手機建單頁不含舊 4-combo", mob_no_old)
    check("①下拉只兩個品項選項(片/盒)", mhtml.count('value="LOOSE"') >= 1 and mhtml.count('value="BOX"') >= 1)

    # ---- ② 盒×3 ----
    r = client.post("/orders/new", data={
        "item_product_id": str(pid), "item_combo_code": "BOX",
        "item_qty": "3", "item_amount": "900",
    }, follow_redirects=True)
    db = get_session()
    check("②盒×3 盒裝 normal -3", _bal(db, pid, "boxed") == box0 - 3,
          f"now={_bal(db, pid, 'boxed')} expect={box0-3}")
    check("②盒×3 裸片不動", _bal(db, pid, "loose_piece") == loose0)
    o = db.query(Order).order_by(Order.id.desc()).first()
    it = db.query(OrderItem).filter_by(order_id=o.id).first()
    check("②盒×3 明細 combo_code=BOX", it.combo_code == "BOX", f"got={it.combo_code}")
    check("②盒×3 金額正確(實收 900)", float(it.subtotal) == 900.0, f"got={it.subtotal}")
    mv = (db.query(InventoryMovement)
          .filter_by(inventory_pool="boxed", movement_type="SALE", ref_id=str(o.id))
          .first())
    check("②盒×3 寫 SALE movement(boxed -3)", mv is not None and mv.qty_delta == -3,
          f"delta={mv.qty_delta if mv else None}")
    box1 = _bal(db, pid, "boxed")
    db.close()

    # ---- 片×5 ----
    r = client.post("/orders/new", data={
        "item_product_id": str(pid), "item_combo_code": "LOOSE",
        "item_qty": "5", "item_amount": "250",
    }, follow_redirects=True)
    db = get_session()
    check("②片×5 裸片 normal -5", _bal(db, pid, "loose_piece") == loose0 - 5,
          f"now={_bal(db, pid, 'loose_piece')} expect={loose0-5}")
    check("②片×5 盒裝不動", _bal(db, pid, "boxed") == box1)
    loose1 = _bal(db, pid, "loose_piece")
    db.close()

    # ---- ③ 多列(盒×2 + 片×4) 同交易扣兩池 ----
    r = client.post("/orders/new", data={
        "item_product_id": [str(pid), str(pid)],
        "item_combo_code": ["BOX", "LOOSE"],
        "item_qty": ["2", "4"],
        "item_amount": ["600", "200"],
    }, follow_redirects=True)
    db = get_session()
    check("③多列 盒裝 -2", _bal(db, pid, "boxed") == box1 - 2,
          f"now={_bal(db, pid, 'boxed')} expect={box1-2}")
    check("③多列 裸片 -4", _bal(db, pid, "loose_piece") == loose1 - 4,
          f"now={_bal(db, pid, 'loose_piece')} expect={loose1-4}")
    box2 = _bal(db, pid, "boxed")
    loose2 = _bal(db, pid, "loose_piece")
    mv_count_before = db.query(InventoryMovement).count()
    db.close()

    # ---- ③ 多列任一缺貨整張 rollback（盒×1 OK + 片×999999 缺貨）----
    r = client.post("/orders/new", data={
        "item_product_id": [str(pid), str(pid)],
        "item_combo_code": ["BOX", "LOOSE"],
        "item_qty": ["1", "999999"],
        "item_amount": ["300", "999"],
    }, follow_redirects=True)
    txt = r.get_data(as_text=True)
    db = get_session()
    check("③缺貨整張 rollback 盒裝不變", _bal(db, pid, "boxed") == box2,
          f"now={_bal(db, pid, 'boxed')} expect={box2}")
    check("③缺貨整張 rollback 裸片不變", _bal(db, pid, "loose_piece") == loose2)
    check("③缺貨不留痕(movement 數不增)",
          db.query(InventoryMovement).count() == mv_count_before,
          f"now={db.query(InventoryMovement).count()} before={mv_count_before}")
    check("③缺貨頁面有提示", "庫存不足" in txt)
    db.close()

    # ---- ④ 扣不到 reserved（預留池不動）----
    db = get_session()
    check("④盒裝 reserved 全程未動", _bal(db, pid, "boxed", "reserved") == box_res0,
          f"now={_bal(db, pid, 'boxed', 'reserved')} expect={box_res0}")
    # 確認沒有任何 SALE movement 碰到 reserved
    res_sale = (db.query(InventoryMovement)
                .filter(InventoryMovement.movement_type == "SALE")
                .filter((InventoryMovement.stock_category_from == "reserved")
                        | (InventoryMovement.stock_category_to == "reserved"))
                .count())
    check("④無任何 SALE 觸及 reserved", res_sale == 0, f"count={res_sale}")
    db.close()

    # ---- ⑤ 顯示 片/盒 中文 ----
    db = get_session()
    o_box = (db.query(OrderItem).filter_by(combo_code="BOX").first())
    oid = o_box.order_id
    db.close()
    det = client.get(f"/orders/{oid}").get_data(as_text=True)
    check("⑤訂單明細顯示『盒』中文", "盒" in det)
    rank = client.get("/reports/sales").get_data(as_text=True)
    check("⑤銷售排行顯示 片/盒", ("片" in rank and "盒" in rank))
    xlsx = client.get("/reports/export.xlsx")
    check("⑤Excel 匯出成功(200)", xlsx.status_code == 200)
    # 驗 Excel 訂單分頁/銷售分頁含片/盒
    import io
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(xlsx.data))
    sales_ws = wb["銷售報表"]
    sales_text = " ".join(
        str(v) for row in sales_ws.iter_rows(values_only=True) for v in row
        if v is not None
    )
    check("⑤Excel 銷售分頁含 片/盒", ("片" in sales_text and "盒" in sales_text),
          sales_text[:120])

    # ---- ⑥ 其他功能沒壞 ----
    check("⑥ROI 頁可開(200)", client.get("/reports/roi").status_code == 200)
    check("⑥支出管理頁可開(200)", client.get("/reports/expenses").status_code == 200)
    check("⑥庫存報表頁可開(200)", client.get("/reports/inventory").status_code == 200)
    check("⑥手機庫存頁可開(200)", client.get("/m/inventory").status_code == 200)
    check("⑥訂單列表可開(200)", client.get("/orders/").status_code == 200)
    check("⑥手機訂單列表可開(200)", client.get("/m/orders").status_code == 200)
    # 權限：viewer 不可建單（403 或被導去登入）
    vc = app.test_client()
    login(vc, "viewer", "viewer123")
    vr = vc.get("/orders/new")
    check("⑥權限：viewer 不可進建單頁(403)", vr.status_code == 403, f"status={vr.status_code}")

    # ---- 手機建單也跑一發確認扣減正確 ----
    db = get_session()
    box_b = _bal(db, pid, "boxed")
    db.close()
    sc = app.test_client()
    login(sc, "staff", "staff123")
    sc.post("/m/orders/new", data={
        "product_id": str(pid), "combo_code": "BOX", "qty": "1", "amount": "300",
    }, follow_redirects=True)
    db = get_session()
    check("⑥手機建單盒×1 盒裝 -1", _bal(db, pid, "boxed") == box_b - 1,
          f"now={_bal(db, pid, 'boxed')} expect={box_b-1}")
    db.close()

    print("\n==== 結果 ====")
    print(f"PASS={len(PASS)}  FAIL={len(FAIL)}")
    if FAIL:
        print("FAILED:")
        for f in FAIL:
            print("  - " + f)
        print("\n總結：FAIL")
        return 1
    print("\n總結：PASS")
    return 0


if __name__ == "__main__":
    sys.exit(main())
