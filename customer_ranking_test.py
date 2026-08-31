# -*- coding: utf-8 -*-
"""CR-3 驗收：客戶銷售排行（reports._customer_ranking / /reports/customers / 客戶列表欄 / Excel 分頁）— 2026-08-31。

throwaway sqlite 放 %TEMP%，不污染交付庫；env 在 import 任何 app 模組前先設定。
樣本：3 客戶 + 5 單 → 甲 2 單（1 paid 1000、1 refunded 5000）、乙 1 單 3000、丙 1 單作廢 9000、未綁客戶 1 單 700
驗證項目：
  A _customer_ranking：金額 desc（乙 3000 > 甲 1000）、refunded 不計金額但計訂單數、voided 完全排除、
    未綁定客戶另列最後、日期篩選（from/to 含當日）
  B /reports/customers：owner 200 含排行與「未綁定客戶」；staff 403；accounting 200；?from=&to= 篩選
  C /customers/ 列表：owner 見 訂單數 / 累計金額 / 最近購買；?sort=rank 乙排在甲前；staff 不見累計金額欄
  D Excel export 含「客戶排行」分頁，內容含乙 3000 與未綁定客戶
"""
import os
import sys
import tempfile
from datetime import datetime
from io import BytesIO

try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except Exception:
    pass

_TMP_DB = os.path.join(tempfile.gettempdir(), "flora_court_custrank_test.db")
if os.path.exists(_TMP_DB):
    os.remove(_TMP_DB)
os.environ["DATABASE_URL"] = "sqlite:///" + _TMP_DB.replace("\\", "/")
os.environ["ENABLE_MOBILE"] = "true"
os.environ["SECRET_KEY"] = "test-secret"

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import init_db  # noqa: E402
from db import get_session, Order, OrderItem, Product, Customer  # noqa: E402
from blueprints.reports import _customer_ranking, UNBOUND_CUSTOMER_LABEL  # noqa: E402
from app import create_app  # noqa: E402

PASS = []
FAIL = []


def check(name, cond, detail=""):
    (PASS if cond else FAIL).append(name)
    print(f"[{'PASS' if cond else 'FAIL'}] {name}" + (f" — {detail}" if detail else ""))


def login(client, username, password):
    client.get("/logout")
    return client.post("/login", data={"username": username, "password": password},
                       follow_redirects=True)


def main():
    print(f"=== throwaway DB: {_TMP_DB} ===")
    init_db.create_all()
    db0 = get_session()
    init_db.seed(db0)
    db0.close()

    db = get_session()
    mask = db.query(Product).filter_by(sku=init_db.MASK_SKU).first()
    ca = Customer(name="排行甲", phone="0911000001")
    cb = Customer(name="排行乙", phone="0911000002")
    cc = Customer(name="排行丙", phone="0911000003")
    db.add_all([ca, cb, cc])
    db.flush()

    def mk(no, cust_id, created, pay, amount, voided=False):
        o = Order(order_no=no, customer_id=cust_id, recipient_name="收件人",
                  total_amount=amount, shipping_fee=60, payment_status=pay,
                  shipping_status="pending", created_at=created)
        if voided:
            o.voided_at = datetime(2026, 8, 30, 12, 0, 0)
            o.void_reason = "測試作廢"
        db.add(o)
        db.flush()
        db.add(OrderItem(order_id=o.id, product_id=mask.id, combo_code="BOX",
                         qty=1, unit_price=amount, subtotal=amount))

    mk("RK-A1", ca.id, datetime(2026, 7, 5, 10, 0), "paid", 1000)
    mk("RK-A2", ca.id, datetime(2026, 8, 15, 10, 0), "refunded", 5000)   # 不計金額、計單數
    mk("RK-B1", cb.id, datetime(2026, 8, 10, 10, 0), "paid", 3000)
    mk("RK-C1", cc.id, datetime(2026, 8, 12, 10, 0), "paid", 9000, voided=True)  # 完全排除
    mk("RK-N1", None, datetime(2026, 8, 20, 10, 0), "unpaid", 700)      # 未綁定客戶
    db.commit()
    ida, idb, idc = ca.id, cb.id, cc.id
    db.close()

    # =====================================================================
    # A：_customer_ranking 純函式
    # =====================================================================
    db = get_session()
    rk = _customer_ranking(db)
    by_id = {r["customer_id"]: r for r in rk}
    check("A 排序：乙(3000) 第一、甲(1000) 第二", [r["customer_id"] for r in rk[:2]] == [idb, ida],
          str([(r["name"], r["total_amount"]) for r in rk]))
    check("A 甲：refunded 不計金額（1000）但計訂單數（2）",
          ida in by_id and by_id[ida]["total_amount"] == 1000.0 and by_id[ida]["order_count"] == 2)
    check("A 甲最近購買 = 2026-08-15（含 refunded 單）",
          by_id[ida]["last_order_at"].strftime("%Y-%m-%d") == "2026-08-15")
    check("A 丙：唯一單已作廢 → 不出現在排行", idc not in by_id)
    check("A 未綁定客戶另列、固定最後、金額 700",
          rk[-1]["customer_id"] is None and rk[-1]["name"] == UNBOUND_CUSTOMER_LABEL
          and rk[-1]["total_amount"] == 700.0 and rk[-1]["order_count"] == 1)
    check("A 回傳欄位齊全", all(k in rk[0] for k in
          ("customer_id", "name", "phone", "order_count", "total_amount", "last_order_at")))
    rk2 = _customer_ranking(db, date_from=datetime(2026, 8, 1), date_to=datetime(2026, 8, 10))
    ids2 = [r["customer_id"] for r in rk2]
    check("A 日期篩選 8/1–8/10（含當日）：只剩乙", ids2 == [idb], str(ids2))
    rk3 = _customer_ranking(db, date_from=datetime(2026, 8, 11))
    ids3 = [r["customer_id"] for r in rk3]
    check("A 日期篩選 8/11 起：甲(僅 refunded 單, 金額 0) + 未綁定",
          set(ids3) == {ida, None} and [r for r in rk3 if r["customer_id"] == ida][0]["total_amount"] == 0.0,
          str([(r["name"], r["total_amount"]) for r in rk3]))
    db.close()

    app = create_app()
    app.config["TESTING"] = True
    c = app.test_client()

    # =====================================================================
    # B：/reports/customers 權限與內容
    # =====================================================================
    login(c, "owner", "owner123")
    r = c.get("/reports/customers")
    html = r.get_data(as_text=True)
    check("B owner /reports/customers 200", r.status_code == 200, f"got {r.status_code}")
    check("B 頁面含 乙 / 甲 / 未綁定客戶，且不含 丙",
          "排行乙" in html and "排行甲" in html and UNBOUND_CUSTOMER_LABEL in html and "排行丙" not in html)
    check("B 乙排在甲前", html.index("排行乙") < html.index("排行甲"))
    check("B 金額 $3000 / $1000 / $700 出現", "$3000" in html and "$1000" in html and "$700" in html)
    check("B 合計 = 3000+1000+700 = $4700", "$4700" in html)
    r = c.get("/reports/customers?from=2026-08-01&to=2026-08-10")
    html = r.get_data(as_text=True)
    check("B 日期篩選頁：只剩乙", "排行乙" in html and "排行甲" not in html and UNBOUND_CUSTOMER_LABEL not in html)
    r = c.get("/reports/customers?from=bad-date")
    check("B 錯誤日期不崩（200 + 提示）", r.status_code == 200 and "日期格式" in r.get_data(as_text=True))
    idx = c.get("/reports/").get_data(as_text=True)
    check("B 報表首頁有客戶排行入口", "/reports/customers" in idx)
    check("B 導覽列有客戶排行（owner）", "客戶排行" in idx)

    login(c, "staff", "staff123")
    r = c.get("/reports/customers")
    check("B staff /reports/customers 403", r.status_code == 403, f"got {r.status_code}")
    check("B staff 導覽列無客戶排行入口", "/reports/customers" not in c.get("/customers/").get_data(as_text=True))
    login(c, "accounting", "accounting123")
    check("B accounting /reports/customers 200", c.get("/reports/customers").status_code == 200)
    login(c, "viewer", "viewer123")
    check("B viewer /reports/customers 403", c.get("/reports/customers").status_code == 403)

    # =====================================================================
    # C：/customers/ 列表欄位 + sort=rank
    # =====================================================================
    login(c, "owner", "owner123")
    html = c.get("/customers/").get_data(as_text=True)
    check("C 列表有 訂單數 / 累計金額 / 最近購買 欄", "訂單數" in html and "累計金額" in html and "最近購買" in html)
    check("C 列表含甲 $1000、乙 $3000、丙 $0", "$1000" in html and "$3000" in html and "$0" in html)
    check("C 預設排序：丙(最新建立) 在最前", html.index("排行丙") < html.index("排行乙") < html.index("排行甲"))
    html = c.get("/customers/?sort=rank").get_data(as_text=True)
    check("C ?sort=rank：乙 > 甲 > 丙", html.index("排行乙") < html.index("排行甲") < html.index("排行丙"))
    check("C 列表最近購買日 2026-08-15（甲）", "2026-08-15" in html)
    login(c, "staff", "staff123")
    html = c.get("/customers/?sort=rank").get_data(as_text=True)
    check("C staff 列表不見累計金額欄", "累計金額" not in html and "$3000" not in html)
    check("C staff 仍見訂單數與最近購買", "訂單數" in html and "2026-08-15" in html)

    # =====================================================================
    # D：Excel 分頁
    # =====================================================================
    login(c, "owner", "owner123")
    resp = c.get("/reports/export.xlsx")
    from openpyxl import load_workbook
    wb = load_workbook(BytesIO(resp.get_data()))
    check("D Excel 分頁含「客戶排行」且順序在銷售報表之後",
          "客戶排行" in wb.sheetnames
          and wb.sheetnames.index("客戶排行") == wb.sheetnames.index("銷售報表") + 1, str(wb.sheetnames))  # CR-8 後另有「操作紀錄」在最後
    ws = wb["客戶排行"]
    rows = [[cell.value for cell in row] for row in ws.iter_rows()]
    check("D 客戶排行表頭", rows[0] == ["名次", "客戶", "電話", "訂單數", "累計金額", "最近購買"], str(rows[0]))
    check("D 第 1 名 = 乙 3000", rows[1][0] == 1 and rows[1][1] == "排行乙" and rows[1][4] == 3000.0, str(rows[1]))
    check("D 最後一列 = 未綁定客戶（無名次）", rows[-1][1] == UNBOUND_CUSTOMER_LABEL and rows[-1][0] in ("", None),
          str(rows[-1]))
    check("D 丙不在分頁", not any(r[1] == "排行丙" for r in rows[1:]))

    print("\n=== 結果 ===")
    print(f"PASS: {len(PASS)}  FAIL: {len(FAIL)}")
    if FAIL:
        print("FAILED:")
        for f in FAIL:
            print("  - " + f)
        sys.exit(1)
    print("ALL GREEN")


if __name__ == "__main__":
    main()
